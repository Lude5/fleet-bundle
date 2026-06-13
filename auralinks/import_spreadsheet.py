"""Build auralinks' catalogue from the client's spreadsheet (ASMRHauler LITBUY xlsx).

The spreadsheet is the source of truth: every product hyperlink in every tab is
extracted, the Weidian/Taobao/1688 source id is parsed out of whatever agent
wrapper the sheet used (CNFans, Litbuy, raw links...), and each item is stored
with a KakoBuy-format buy URL carrying THIS client's affcode (the engine's
native format — the agent picker rebuilds per-agent links with his other codes).
The sheet author's own referral codes are deliberately discarded.

Images aren't in the sheet (just embedded thumbnails on a few rows), so images
are resolved from the previous catalogue by source itemID where the same item
already existed; everything else gets the placeholder until enrichment.

Usage:  python import_spreadsheet.py [--write]
Without --write it's a dry run: prints per-tab stats + samples, writes nothing.
With --write it replaces static/products.json (old items REMOVED, per client).
"""
import sys, os, re, json, hashlib
from urllib.parse import urlparse, parse_qs, unquote, quote

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = r"C:\Users\kaiom\OneDrive\Pictures\ASMRHauler LITBUY Spreadsheet +8000 Items（副本）.xlsx"
OLD_JSON = os.path.join(HERE, 'static', 'products.json')
OUT = os.path.join(HERE, 'static', 'products.json')
AFFCODE = 'n62s8s'
WRITE = '--write' in sys.argv

# tab name -> category strategy. 'split' uses name keywords inside combined tabs.
TABS = {
    'ASMRHauler CNFans Spreadsheet': {'cat': 'byname', 'trending': False},
    '🔥Trending Now 🔥  2':          {'cat': 'byname', 'trending': True},
    '🔥Trending Now 🔥  ':           {'cat': 'byname', 'trending': True},
    '🐼 Update ⛄':                  {'cat': 'byname', 'trending': False},
    '👞SHOES👞':                     {'cat': 'shoes', 'trending': False},
    '👕T-shirt and shorts🩳':        {'cat': 'split:shirts', 'trending': False},
    '👜 Accessories👜':              {'cat': 'split:accessories', 'trending': False},
    '🎧Electronic products🎧':       {'cat': 'tech', 'trending': False},
    '👠 Womens Spreadsheet 👗':      {'cat': 'womens', 'trending': False},
    '🥼Hoodies and Pants👖':         {'cat': 'split:hoodies', 'trending': False},
    '🧥Coats and Jackets🧥':         {'cat': 'jackets', 'trending': False},
    '👠 Womens Spreadsheet 👗1':     {'cat': 'womens', 'trending': False},
}

NUM_PREFIX = re.compile(r'^\s*\d+\s*[、,．.]\s*')
PRICE_RE = re.compile(r'(\d+(?:[.,]\d+)?)\s*\$')

def clean_name(s):
    s = str(s or '').replace('\r', ' ').replace('\n', ' ')
    s = NUM_PREFIX.sub('', s)
    s = re.sub(r'[🔥⚡✨❄️☃️🐼⛄]', '', s)
    return re.sub(r'\s+', ' ', s).strip(' -–—|')

def source_from_link(target):
    """Return (platform, item_id) from any agent/product wrapper, else (None, None)."""
    if not target:
        return None, None
    t = unquote(str(target))
    # direct or nested weidian / taobao / 1688
    m = re.search(r'weidian\.com/item\.html\?[^\s"]*?itemI[dD]=(\d+)', t)
    if m: return 'weidian', m.group(1)
    m = re.search(r'item\.taobao\.com/item\.htm\?[^\s"]*?id=(\d+)', t)
    if m: return 'taobao', m.group(1)
    m = re.search(r'detail\.1688\.com/offer/(\d+)\.html', t)
    if m: return '1688', m.group(1)
    # litbuy path style: litbuy.com/product/weidian/7718385571?inviteCode=...
    m = re.search(r'/product/(weidian|wd|taobao|tb|ali_?1688|1688)/(\d+)', t, re.I)
    if m:
        plat = m.group(1).lower()
        plat = 'weidian' if plat in ('weidian', 'wd') else ('taobao' if plat in ('taobao', 'tb') else '1688')
        return plat, m.group(2)
    # cnfans / similar: ?shop_type=weidian&id=123  (also platform= and pid= variants)
    q = parse_qs(urlparse(t).query)
    st = (q.get('shop_type') or q.get('platform') or q.get('channel') or [''])[0].lower()
    pid = (q.get('id') or q.get('pid') or q.get('itemID') or q.get('itemId') or [''])[0]
    if pid.isdigit():
        if 'weidian' in st or 'wd' == st: return 'weidian', pid
        if 'taobao' in st or 'tb' == st:  return 'taobao', pid
        if '1688' in st or 'ali' in st:   return '1688', pid
        if st == '' and 'cnfans' in t:    return 'weidian', pid   # cnfans default
    return None, None

def buy_url(platform, item_id):
    if platform == 'weidian':
        inner = f'https://weidian.com/item.html?itemID={item_id}'
    elif platform == 'taobao':
        inner = f'https://item.taobao.com/item.htm?id={item_id}'
    else:
        inner = f'https://detail.1688.com/offer/{item_id}.html'
    return f'https://www.kakobuy.com/item/details?url={quote(inner, safe="")}&affcode={AFFCODE}'

def guess_cat(name):
    n = name.lower()
    rules = [
        ('shoes', r'\b(shoe|sneaker|jordan|dunk|yeezy|air force|air max|new balance|nb \d|asics|samba|campus|gazelle|spezial|slipper|slide|sandal|loafer|boot|birkenstock|crocs|foam|trainer|b22|b30|b23|b25|mcqueen|tn\b|shox|p6000|vomero|uptempo|spray|kobe|protro|lebron|kyrie|kd \d|zoom|blazer|cortez|foamposite|penny|mag\b|sb dunk|aj\d|af1|triple s|track\b|runner|cloudmonster|on cloud|salomon|merrell|timberland|ugg|heels?\b|pumps\b|mary jane)'),
        ('shorts', r'\bshorts\b'),
        ('hoodies', r'\b(hoodie|sweater|sweatshirt|crewneck|zip(per)? hoodie|knit)\b'),
        ('pants', r'\b(pants|jeans|trousers|sweatpants|cargos?|joggers?)\b'),
        ('jackets', r'\b(jacket|coat|puffer|vest|gilet|windbreaker|parka|down\b)'),
        ('tracksuits', r'\b(set|tracksuit)\b'),
        ('bags', r'\b(bag|backpack|tote|duffle|pouch|wallet|cardholder|keepall)\b'),
        ('headwear', r'\b(hat|cap|beanie|balaclava)\b'),
        ('tech', r'\b(airpods?|headphone|speaker|jbl|dyson|earbuds?|watch ultra|charger|stanley)\b'),
        ('accessories', r'\b(belt|sunglass|glasses|watch|whatch|bracelet|necklace|ring|scarf|sock|underwear|keychain|jewel|tie\b|gloves?)\b'),
        ('shirts', r'\b(tee|t-?shirt|shirt|polo|jersey|long ?sleeve|longtee|long tee|top\b|blouse)\b'),
    ]
    for cat, pat in rules:
        if re.search(pat, n):
            return cat
    return 'accessories'

def cat_for(tabcfg, name):
    c = tabcfg['cat']
    if c == 'byname':
        return guess_cat(name)
    if c.startswith('split:'):
        g = guess_cat(name)
        default = c.split(':', 1)[1]
        # within a combined tab, trust strong name hits; else the tab default
        return g if g != 'accessories' or default == 'accessories' else default
    return c

def harvest(ws, tabname, cfg):
    """Find every product hyperlink; the item name/price live near the link cell."""
    items = []
    grid = {}
    for row in ws.iter_rows():
        for c in row:
            grid[(c.row, c.column)] = c
    for (r, col), c in list(grid.items()):
        hl = c.hyperlink
        target = getattr(hl, 'target', None) if hl else None
        plat, pid = source_from_link(target)
        if not pid:
            continue
        # NAME: master-list layout (link col E, name col B) or grid layout (name
        # immediately left of the LINK cell)
        name_cell = grid.get((r, col - 3)) if str(c.value or '').strip().upper().startswith('CNFANS') else None
        if not (name_cell and name_cell.value):
            name_cell = grid.get((r, col - 1))
        name = clean_name(name_cell.value if name_cell else '')
        if not name or len(name) < 3 or name.strip().upper() in ('LINK', 'CNFANS LINK', 'QC'):
            continue
        # PRICE: master layout -> col C (link col - 2); grid -> usd at col+2 ('44.62$' text or number)
        usd = None
        for cand in (grid.get((r, col - 2)), grid.get((r, col + 2)), grid.get((r, col + 1))):
            if cand is None or cand.value is None: continue
            v = cand.value
            if isinstance(v, (int, float)) and 0.5 < float(v) < 5000:
                # grid CNY col is also numeric; prefer cells with $ text, else
                # treat: master tab numeric col C IS usd. Disambiguate by tab.
                usd = float(v); break
            m = PRICE_RE.search(str(v))
            if m:
                usd = float(m.group(1).replace(',', '.')); break
        if usd is None:
            continue
        # grid tabs: col+1 = CNY (number), col+2 = '44.62$' (usd text). The loop
        # above prefers col-2 (master), then col+2 ($ text matched first when CNY
        # numeric also matched col+2? no: order col-2, col+2, col+1). For grid,
        # col-2 is usually empty/name of prior block; col+2 has the $ text.
        items.append({
            'platform': plat, 'pid': pid, 'name': name, 'usd': round(usd, 2),
            'tab': tabname, 'trending': cfg['trending'],
        })
    return items

ENRICH_CACHE = os.path.join(HERE, 'static', 'enrich_cache.json')

_IMG_JUNK = ('wx_default_headimg', 'hz_img_', 'icon-', '/avatar', 'login_', 'wd_logo', 'common-')
_IMG_RE = re.compile(r'https?://[a-z]+\.geilicdn\.com/(?:pcitem|open)[^"\'\\\s&]+?\.(?:jpg|jpeg|png|webp)[^"\'\\\s&]*')

def fetch_weidian_images(pid, tries=3):
    """Product gallery images. Primary: the public item PAGE (robust, holds the
    full gallery). Fallback: the SKU API (per-colour images). Returns a list, or
    None on hard failure — failures must NEVER be cached as 'no image'."""
    import urllib.request, time, random
    UA = {'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1'}
    for attempt in range(tries):
        try:
            req = urllib.request.Request(f'https://weidian.com/item.html?itemID={pid}', headers=UA)
            with urllib.request.urlopen(req, timeout=12) as r:
                html = r.read().decode('utf-8', 'ignore')
            imgs, seen = [], set()
            for u in _IMG_RE.findall(html):
                if any(j in u for j in _IMG_JUNK) or u in seen:
                    continue
                seen.add(u); imgs.append(u)
                if len(imgs) >= 10:
                    break
            if imgs:
                return imgs
            # page had no gallery -> one shot at the SKU API
            p = quote(json.dumps({"itemId": pid}, separators=(',', ':')))
            req2 = urllib.request.Request(
                f"https://thor.weidian.com/detail/getItemSkuInfo/1.0?param={p}",
                headers={**UA, 'Referer': f'https://shop.weidian.com/item.html?itemID={pid}', 'Origin': 'https://shop.weidian.com'})
            with urllib.request.urlopen(req2, timeout=10) as r:
                d = json.loads(r.read().decode('utf-8'))
            for attr in ((d.get('result') or {}).get('attrList') or []):
                for av in (attr.get('attrValues') or []):
                    img = av.get('img')
                    if img and img.startswith('http') and img not in imgs:
                        imgs.append(img)
            return imgs[:10]   # may be genuinely [] after both sources answered
        except Exception:
            time.sleep(0.7 * (attempt + 1) + random.random() * 0.5)
    return None


def enrich_images(records):
    """Fill image/images for weidian items via the SKU API (cached on disk)."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    cache = {}
    if os.path.exists(ENRICH_CACHE):
        try: cache = json.load(open(ENRICH_CACHE, encoding='utf-8'))
        except Exception: cache = {}
    # only NON-EMPTY cache entries are trusted; old empty entries may be cached
    # rate-limit failures from a previous run, so they get retried
    cache = {k: v for k, v in cache.items() if v}
    todo = [r for r in records if not r['image'] and r['_platform'] == 'weidian']
    cache_only = bool(os.environ.get('CACHE_ONLY'))
    print(f'enriching images for {len(todo)} weidian items ({len(cache)} cached, cache_only={cache_only})...')
    done = failed = 0
    with ThreadPoolExecutor(max_workers=10) as ex:
        futs = {}
        for r in todo:
            if r['_pid'] in cache:
                imgs = cache[r['_pid']]
                r['image'], r['images'] = imgs[0], imgs
                continue
            if cache_only:
                continue   # the detached crawler owns the network; don't double-hit
            futs[ex.submit(fetch_weidian_images, r['_pid'])] = r
        for fut in as_completed(futs):
            r = futs[fut]
            imgs = fut.result()
            if imgs:                      # success with images
                cache[r['_pid']] = imgs
                r['image'], r['images'] = imgs[0], imgs
            elif imgs is None:            # hard failure -> NOT cached, retryable
                failed += 1
            done += 1
            if done % 400 == 0:
                print(f'  ...{done}/{len(futs)} (failed so far: {failed})')
                json.dump(cache, open(ENRICH_CACHE, 'w', encoding='utf-8'))
    json.dump(cache, open(ENRICH_CACHE, 'w', encoding='utf-8'))
    have = sum(1 for r in records if r['image'])
    print(f'images: {have}/{len(records)} have a real photo | hard failures this pass: {failed}')


def build_sheet_images(path):
    """Map (platform, pid, canon_name) -> the row's =IMAGE() url, and (platform,pid) ->
    first url. MUST read the FORMULA layer (data_only=False): an =IMAGE() cell renders a
    picture and has NO cached value, so data_only=True returns blank for it — which is
    why these per-product photos were missed and items fell back to itemID-matched ones."""
    wbf = openpyxl.load_workbook(path, data_only=False)
    # exclude commas (incl. fullwidth ，) so `=IMAGE("url"，2)` mode params don't get
    # glued onto the URL (that 404'd a handful of otherwise-valid postimg photos).
    img_re = re.compile(r'https?://[^\s"“”\)\'，,]+')
    by_key, by_pid = {}, {}
    for ws in wbf.worksheets:
        grid = {(c.row, c.column): c for row in ws.iter_rows() for c in row}
        # link columns per row — the sheet is a GRID: up to 4 products side-by-side
        # per row, each in its own block [name, link, price, $, IMAGE]. The image
        # MUST be read from THIS product's block, not the first IMAGE() in the row
        # (that bug stamped one photo onto every product sharing a row).
        links_by_row = {}
        for (r, col), c in grid.items():
            _, p = source_from_link(getattr(c.hyperlink, 'target', None) if c.hyperlink else None)
            if p:
                links_by_row.setdefault(r, []).append(col)
        for (r, col), c in list(grid.items()):
            plat, pid = source_from_link(getattr(c.hyperlink, 'target', None) if c.hyperlink else None)
            if not pid:
                continue
            is_cnfans = str(c.value or '').strip().upper().startswith('CNFANS')
            ncell = grid.get((r, col - 3)) if is_cnfans else None
            if not (ncell and ncell.value):
                ncell = grid.get((r, col - 1))
            name = clean_name(ncell.value if ncell else '')
            if not name or len(name) < 3:
                continue
            # this product's block: CNFANS layout puts the IMAGE to the LEFT of the
            # link (col 1); grid layout puts it to the RIGHT, before the next link.
            if is_cnfans:
                scan = range(col - 1, 0, -1)
            else:
                nxt = min([x for x in links_by_row.get(r, []) if x > col], default=ws.max_column + 1)
                scan = range(col + 1, nxt)
            img = ''
            for cc in scan:
                v = ws.cell(r, cc).value
                if isinstance(v, str) and 'IMAGE(' in v.upper():
                    m = img_re.search(v)
                    if m:
                        img = m.group(0)
                        break
            if 'load-failure' in img.lower():   # curator's broken-image marker -> placeholder
                img = ''
            if not img:
                continue
            by_key.setdefault((plat, pid, re.sub(r'\s+', ' ', name).strip().lower()), img)
            by_pid.setdefault((plat, pid), img)
    print(f'sheet =IMAGE() photos: {len(by_pid)} items ({len(by_key)} (id,name) keys)')
    return by_key, by_pid


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=False, data_only=True)
    all_items, per_tab = [], {}
    for tabname, cfg in TABS.items():
        if tabname not in wb.sheetnames:
            print(f'!! tab missing: {tabname}'); continue
        got = harvest(wb[tabname], tabname, cfg)
        per_tab[tabname] = len(got)
        all_items.extend(got)

    # ------------------------------------------------------------------
    # Dedupe. The sheet attaches ONE listing link to MULTIPLE products
    # (sellers' multi-product listings: e.g. one link named "Fendi Shorts",
    # "Adidas Slippers" AND "Goyard Cardholder"). Collapsing by pid alone
    # mixed one entry's name/price with another's listing — so the key is
    # (pid, canonical-name): every distinct product entry survives with its
    # own name + price. Within a pid, only true SYNONYMS are merged
    # ("LV Belt" / "Louis Vuitton Belt" / same first word or substring),
    # keeping the most descriptive name and preferring a category-tab record.
    # ------------------------------------------------------------------
    def specificity(it):
        return 0 if TABS[it['tab']]['cat'] in ('byname',) else 1

    def canon(n):
        return re.sub(r'\s+', ' ', n).strip().lower()

    _SYN = {'louis vuitton': 'lv', 'christian louboutin': 'louboutin'}
    def syn_form(n):
        s = canon(n)
        for long, short in _SYN.items():
            s = s.replace(long, short)
        return s

    def similar(a, b):
        a, b = syn_form(a), syn_form(b)
        if a == b or a in b or b in a:
            return True
        return a.split()[0] == b.split()[0] and len(set(a.split()) & set(b.split())) >= 2

    trending_ids = set()
    groups = {}   # pid -> list of variant dicts
    for it in all_items:
        pid_key = (it['platform'], it['pid'])
        if it['trending']:
            trending_ids.add(pid_key)
        variants = groups.setdefault(pid_key, [])
        for v in variants:
            if similar(v['name'], it['name']):
                # synonym of an existing entry: keep the better record
                if (specificity(it), len(it['name'])) > (specificity(v), len(v['name'])):
                    v.update(it)
                break
        else:
            variants.append(dict(it))

    best = {}
    for pid_key, variants in groups.items():
        for v in variants:
            best[(pid_key[0], pid_key[1], canon(v['name']))] = v
    n_multi = sum(1 for vs in groups.values() if len(vs) > 1)
    print(f'listings with multiple distinct products kept separate: {n_multi}')

    # old catalogue: image + extra images by source pid (any platform)
    old_by_pid = {}
    try:
        old = json.load(open(OLD_JSON, encoding='utf-8'))
        old_items = old if isinstance(old, list) else old.get('products', [])
        for p in old_items:
            m = re.search(r'itemID(?:%3D|=)(\d+)', p.get('url') or '', re.I)
            if m:
                old_by_pid[m.group(1)] = p
    except Exception as e:
        print('old catalogue not readable:', e)

    # neutral tile for items whose photo hasn't been fetched yet (empty <img src>
    # doesn't reliably fire onerror, so the JS fallback can't catch it)
    PLACEHOLDER = ('data:image/svg+xml;utf8,' +
        '%3Csvg xmlns=%22http://www.w3.org/2000/svg%22 viewBox=%220 0 100 100%22%3E'
        '%3Crect width=%22100%22 height=%22100%22 fill=%22%23f2f2f4%22/%3E'
        '%3Cg fill=%22none%22 stroke=%22%23c7c7cf%22 stroke-width=%223%22%3E'
        '%3Crect x=%2230%22 y=%2234%22 width=%2240%22 height=%2232%22 rx=%223%22/%3E'
        '%3Ccircle cx=%2241%22 cy=%2246%22 r=%224%22/%3E'
        '%3Cpath d=%22M34 62l12-12 8 8 8-9 8 13%22/%3E%3C/g%3E%3C/svg%3E')

    sheet_key, sheet_pid = build_sheet_images(XLSX)

    out, with_img = [], 0
    for (plat, pid, cname), it in best.items():
        cat = cat_for(TABS[it['tab']], it['name'])
        # Images come from the spreadsheet's own =IMAGE() cells ONLY (client
        # decision) — never from the listing/Weidian. No sheet photo -> the
        # neutral placeholder tile.
        sheet_img = sheet_key.get((plat, pid, cname)) or sheet_pid.get((plat, pid)) or ''
        image = sheet_img
        images = [sheet_img] if sheet_img else []
        if image: with_img += 1
        # id hashes pid + canonical name so multi-product listings keep
        # one catalogue entry PER product
        pid_hash = 'p' + hashlib.md5(f'{plat}:{pid}:{cname}'.encode()).hexdigest()[:10]
        rec = {
            'id': pid_hash,
            'name': it['name'],
            'category': cat,
            'price': it['usd'],
            'price_numeric': it['usd'],
            'url': buy_url(plat, pid),
            'image': image,
            'seller': '', 'batch': '', 'retail_price': '',
            'tags': 'trending' if (plat, pid) in trending_ids else '',
            'images': images,
            '_platform': plat, '_pid': pid,
        }
        out.append(rec)
    for r in out:
        r.pop('_platform', None); r.pop('_pid', None)

    print('=== per-tab raw link hits ===')
    for t, n in per_tab.items():
        print(f'  {t!r}: {n}')
    print(f'total raw: {len(all_items)} | unique products: {len(out)} | with image from old catalogue: {with_img}')
    from collections import Counter
    print('categories:', dict(Counter(r["category"] for r in out).most_common()))
    print('trending  :', sum(1 for r in out if r['tags'] == 'trending'))
    print('\n=== samples ===')
    for r in out[:6] + out[len(out)//2:len(out)//2+3]:
        print(f"  {r['name'][:46]!r:50} ${r['price']:<7} {r['category']:<12} img={'Y' if r['image'] else '-'} {r['url'][:70]}")

    if WRITE:
        for r in out:
            if not r['image']:
                r['image'] = PLACEHOLDER
                r['images'] = []
        json.dump(out, open(OUT, 'w', encoding='utf-8'), ensure_ascii=False, indent=0)
        print(f'\nWROTE {OUT} ({len(out)} products) — old catalogue replaced')
    else:
        print('\nDRY RUN (pass --write to replace static/products.json)')

if __name__ == '__main__':
    main()
