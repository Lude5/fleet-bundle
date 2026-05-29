import os
import json
import secrets
from flask import Flask, render_template, request, jsonify, redirect, session, url_for, send_file, make_response
from flask_cors import CORS
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
# SECRET_KEY: prefer env var, else persist a random key to disk so it survives
# server restarts (otherwise every restart logs every admin out with a
# confusing 'Unauthorized' toast).
_secret_from_env = os.environ.get('SECRET_KEY')
if _secret_from_env:
    app.secret_key = _secret_from_env
else:
    _data_dir = os.environ.get('ATEGOAT_DATA_DIR') or ('/data' if os.path.exists('/data') else os.path.join(os.path.dirname(__file__), 'data'))
    os.makedirs(_data_dir, exist_ok=True)
    _key_path = os.path.join(_data_dir, '.secret_key')
    if os.path.exists(_key_path):
        with open(_key_path, 'r', encoding='utf-8') as _kf:
            app.secret_key = _kf.read().strip() or secrets.token_hex(32)
    else:
        app.secret_key = secrets.token_hex(32)
        try:
            with open(_key_path, 'w', encoding='utf-8') as _kf:
                _kf.write(app.secret_key)
        except Exception:
            pass
# Keep sessions alive for 24 hours so an admin who logs in once stays in
# across page refreshes / new tabs / browser restarts (matches master_admin).
from datetime import timedelta as _td
app.permanent_session_lifetime = _td(hours=24)
# Belt-and-suspenders cookie flags so the session survives top-level navigation
# and isn't dropped by browser quirks (Lax = sent on same-site nav, not stripped
# on link-click). HttpOnly is on by default in Flask.
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_HTTPONLY'] = True
CORS(app)


# Upgrade every admin request's session to permanent — this catches old sessions
# that were created before the 24h lifetime was set so they stop getting kicked.
@app.before_request
def _persist_admin_session():
    if request.path.startswith('/admin') and session.get('admin_logged_in'):
        session.permanent = True

# === SITE CONFIG (all customizable via env vars) ===
SITE_CONFIG = {
    'name': os.environ.get('SITE_NAME', 'AteGoat'),
    'name_part1': os.environ.get('SITE_NAME_PART1', 'Ate'),
    'name_part2': os.environ.get('SITE_NAME_PART2', 'Goat'),
    'domain': os.environ.get('SITE_DOMAIN', 'ategoat.com'),
    'agent_name': os.environ.get('AGENT_NAME', 'KakoBuy'),
    'agent_domain': os.environ.get('AGENT_DOMAIN', 'kakobuy.com'),
    'agent_signup_url': os.environ.get('AGENT_SIGNUP_URL', 'https://www.kakobuy.com/register?affcode=ategoat'),
    'agent_product_url': os.environ.get('AGENT_PRODUCT_URL', 'https://www.kakobuy.com/item/details'),
    'affiliate_code': os.environ.get('AFFILIATE_CODE', 'ategoat'),
    'brand_color': os.environ.get('BRAND_COLOR', '#0d9488'),
    'brand_color_shadow': os.environ.get('BRAND_COLOR_SHADOW', '#0f766e'),
    'meta_pixel_id': os.environ.get('META_PIXEL_ID', ''),
    'coupon_amount': os.environ.get('COUPON_AMOUNT', '400'),
    'tagline': os.environ.get('TAGLINE', 'A curated catalogue of 5,000+ finds. Updated daily.'),
    'product_count_label': os.environ.get('PRODUCT_COUNT_LABEL', '5,800+'),
    'discord_url': os.environ.get('DISCORD_URL', ''),
}

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'changeme123')
ADMIN_API_TOKEN = os.environ.get('ADMIN_API_TOKEN', '')  # for cross-site master-admin calls
DATA_DIR = os.environ.get('ATEGOAT_DATA_DIR') or ('/data' if os.path.exists('/data') else os.path.join(os.path.dirname(__file__), 'data'))
os.makedirs(DATA_DIR, exist_ok=True)

try:
    try:
        from .database import (
            init_db, get_products, get_product, add_product, add_products_bulk,
            update_product, delete_product, search_products, get_categories, add_category,
            update_category, delete_category, count_products_in_category,
            cache_get, cache_set,
            record_click, get_analytics, backup_database, check_auto_backup,
            set_featured, move_category, reorder_products, get_listing_variants,
            get_top_clicked_products, set_in_stock
        )
    except ImportError:
        from database import (
            init_db, get_products, get_product, add_product, add_products_bulk,
            update_product, delete_product, search_products, get_categories, add_category,
            update_category, delete_category, count_products_in_category,
            cache_get, cache_set,
            record_click, get_analytics, backup_database, check_auto_backup,
            set_featured, move_category, reorder_products, get_listing_variants,
            get_top_clicked_products, set_in_stock
        )
    init_db()
    check_auto_backup()

    # Always seed/refresh from static/products.json when its hash changes.
    # This keeps the DB in sync with whatever was last committed without
    # losing user-added products between identical deploys.
    import hashlib
    products_file = os.path.join(os.path.dirname(__file__), 'static', 'products.json')
    if os.path.exists(products_file):
        with open(products_file, 'rb') as _bf:
            json_hash = hashlib.md5(_bf.read()).hexdigest()
        marker_path = os.path.join(DATA_DIR, '.products_hash')
        seeded_hash = ''
        if os.path.exists(marker_path):
            try:
                seeded_hash = open(marker_path, 'r', encoding='utf-8').read().strip()
            except Exception:
                seeded_hash = ''
        if json_hash != seeded_hash:
            # Wipe + re-seed. INSERT OR REPLACE on its own would leave stale
            # rows that aren't in the new json — wipe makes this idempotent.
            try:
                try:
                    from .database import get_db as _get_db
                except ImportError:
                    from database import get_db as _get_db
                _c = _get_db()
                _c.execute('DELETE FROM products')
                _c.commit()
                _c.close()
            except Exception as _e:
                print(f"Wipe warning: {_e}")
            with open(products_file, 'r', encoding='utf-8') as _f:
                _products = json.load(_f)
            add_products_bulk(_products)
            try:
                with open(marker_path, 'w', encoding='utf-8') as _mf:
                    _mf.write(json_hash)
            except Exception as _e:
                print(f"Marker write warning: {_e}")
            print(f"Re-seeded {len(_products)} products (hash {json_hash[:8]})")
        else:
            print(f"Seed up to date ({get_products() and len(get_products()) or 0} products, hash {json_hash[:8]})")

    if not get_categories():
        CATS = [
            {'slug': 'trending', 'name': 'Trending', 'sort_order': 0},
            {'slug': 'shoes', 'name': 'Shoes', 'sort_order': 1},
            {'slug': 'shirts', 'name': 'Shirts', 'sort_order': 2},
            {'slug': 'shorts', 'name': 'Shorts', 'sort_order': 3},
            {'slug': 'hoodies', 'name': 'Hoodies', 'sort_order': 4},
            {'slug': 'tracksuits', 'name': 'Tracksuits', 'sort_order': 5},
            {'slug': 'pants', 'name': 'Pants', 'sort_order': 6},
            {'slug': 'jackets', 'name': 'Jackets', 'sort_order': 7},
            {'slug': 'headwear', 'name': 'Headwear', 'sort_order': 8},
            {'slug': 'accessories', 'name': 'Accessories', 'sort_order': 9},
            {'slug': 'bags', 'name': 'Bags', 'sort_order': 10},
            {'slug': 'tech', 'name': 'Tech', 'sort_order': 11},
            {'slug': 'womens', 'name': 'Womens', 'sort_order': 12},
        ]
        for c in CATS:
            add_category(c['slug'], c['name'], '', '', c['sort_order'])
        print("Categories seeded")
    else:
        # Ensure newly-introduced categories show up even on DBs that were
        # seeded before they existed. add_category is INSERT OR REPLACE so
        # this is idempotent and won't clobber renames the operator made.
        existing_slugs = {c['slug'] for c in get_categories()}
        for slug, name, order in [('shorts', 'Shorts', 5), ('headwear', 'Headwear', 7), ('tracksuits', 'Tracksuits', 5)]:
            if slug not in existing_slugs:
                add_category(slug, name, '', '', order)
                print(f"Backfilled category: {slug}")
except Exception as e:
    print(f"DB init warning: {e}")


# ============================================================
# Shopping agents — generate per-agent buy URLs for any product
# ============================================================
def _parse_item_url(url):
    """Return (platform, item_id) for a raw seller URL, or (None, None)."""
    import re
    if not url:
        return (None, None)
    m = re.search(r'weidian\.com/item\.html\?[^"\s]*itemID=(\d+)', url, re.I)
    if m: return ('weidian', m.group(1))
    m = re.search(r'taobao\.com/item\.htm\?[^"\s]*id=(\d+)', url, re.I)
    if m: return ('taobao', m.group(1))
    m = re.search(r'1688\.com/offer/(\d+)\.html', url, re.I)
    if m: return ('1688', m.group(1))
    return (None, None)


def _unwrap_agent_url(url):
    """If url is an agent wrapper (kakobuy?url=..., joyagoo?url=..., etc.)
    return the inner seller URL, else return as-is."""
    from urllib.parse import urlparse, parse_qs, unquote
    if not url:
        return ''
    try:
        p = urlparse(url)
        qs = parse_qs(p.query)
        for key in ('url', 'productUrl', 'product_url'):
            if key in qs and qs[key]:
                return unquote(qs[key][0])
    except Exception:
        pass
    return url


# Each agent has a builder function (raw_url, item_id, platform, affcode) -> url
# Builders return None when they can't construct a URL (e.g. agents that need
# an item_id we couldn't parse).
def _b_kakobuy(url, _id, _plat, code):
    from urllib.parse import quote
    if not url: return None
    out = f'https://www.kakobuy.com/item/details?url={quote(url, safe="")}'
    if code: out += f'&affcode={code}'
    return out

def _b_joyagoo(url, _id, _plat, code):
    from urllib.parse import quote
    if not url: return None
    out = f'https://www.joyagoo.com/index/item/index.html?url={quote(url, safe="")}'
    if code: out += f'&affcode={code}'
    return out

def _b_sugargoo(url, _id, _plat, code):
    from urllib.parse import quote
    if not url: return None
    out = f'https://www.sugargoo.com/index/item/index.html?url={quote(url, safe="")}'
    if code: out += f'&shareCode={code}'
    return out

def _b_allchinabuy(url, _id, _plat, code):
    from urllib.parse import quote
    if not url: return None
    out = f'https://www.allchinabuy.com/en/page/buy/?from=search-input&url={quote(url, safe="")}'
    if code: out += f'&partnercode={code}'
    return out

def _b_mulebuy(_url, item_id, platform, code):
    if not item_id or not platform: return None
    out = f'https://mulebuy.com/product/?shop_type={platform}&id={item_id}'
    if code: out += f'&affcode={code}'
    return out

def _b_oopbuy(_url, item_id, platform, code):
    if not item_id or not platform: return None
    out = f'https://oopbuy.com/product/{platform}/{item_id}'
    if code: out += f'?inviteCode={code}'
    return out

def _b_hubbuy(url, _id, _plat, code):
    """HubBuy — successor to AllChinaBuy at hubbuycn.com."""
    from urllib.parse import quote
    if not url: return None
    out = f'https://www.hubbuycn.com/order/buy/?from=search-input&url={quote(url, safe="")}'
    if code: out += f'&partnercode={code}'
    return out

def _b_acbuy(_url, item_id, platform, code):
    if not item_id or not platform: return None
    out = f'https://acbuy.com/product?source={platform}&id={item_id}'
    if code: out += f'&u={code}'
    return out

def _b_litbuy(url, _id, _plat, code):
    from urllib.parse import quote
    if not url: return None
    out = f'https://www.litbuy.com/transmit?url={quote(url, safe="")}'
    if code: out += f'&affcode={code}'
    return out

def _b_hipobuy(url, _id, _plat, code):
    from urllib.parse import quote
    if not url: return None
    out = f'https://www.hipobuy.com/order/buy?url={quote(url, safe="")}'
    if code: out += f'&affcode={code}'
    return out

def _b_usfans(url, _id, _plat, code):
    from urllib.parse import quote
    if not url: return None
    out = f'https://www.usfans.com/product/buy?url={quote(url, safe="")}'
    if code: out += f'&affcode={code}'
    return out

# Signup URLs: only 3 (KakoBuy, Oopbuy, Hipobuy) are verified — pulled from
# ategoat.com's public /promos/list endpoint. The other 9 use plain registration
# URLs with NO referral code. To enable affiliate income for those, drop the
# operator's own codes in via env vars or edit the URLs directly.
AGENTS = [
    {'key': 'kakobuy',     'name': 'KakoBuy',     'build': _b_kakobuy,     'color': '#0d9488', 'domain': 'kakobuy.com',     'signup': 'https://www.kakobuy.com/register/?affcode=bswes',           'coupon': 'Up to $500 shipping credit'},
    {'key': 'oopbuy',      'name': 'Oopbuy',      'build': _b_oopbuy,      'color': '#22c55e', 'domain': 'oopbuy.com',      'signup': 'https://oopbuy.com/register?inviteCode=KRLHFHSGL',          'coupon': 'Up to $300 in coupons'},
    {'key': 'hipobuy',     'name': 'Hipobuy',     'build': _b_hipobuy,     'color': '#14b8a6', 'domain': 'hipobuy.com',     'signup': 'https://hipobuy.com/register?inviteCode=25RXG9B0E',         'coupon': 'Up to $100 in coupons'},
    {'key': 'joyagoo',     'name': 'JoyaGoo',     'build': _b_joyagoo,     'color': '#ef4444', 'domain': 'joyagoo.com',     'signup': 'https://www.joyagoo.com/index/user/register',               'coupon': 'Up to $300 in coupons'},
    {'key': 'sugargoo',    'name': 'Sugargoo',    'build': _b_sugargoo,    'color': '#ec4899', 'domain': 'sugargoo.com',    'signup': 'https://www.sugargoo.com/index/user/register',              'coupon': 'Up to $200 in coupons'},
    {'key': 'hubbuy',      'name': 'HubBuy',      'build': _b_hubbuy,      'color': '#3b82f6', 'domain': 'hubbuycn.com',    'signup': 'https://www.hubbuycn.com/register',                         'coupon': 'Up to $150 in coupons'},
    {'key': 'mulebuy',     'name': 'Mulebuy',     'build': _b_mulebuy,     'color': '#a855f7', 'domain': 'mulebuy.com',     'signup': 'https://mulebuy.com/register',                             'coupon': 'Up to $200 in coupons'},
    {'key': 'acbuy',       'name': 'ACBuy',       'build': _b_acbuy,       'color': '#8b5cf6', 'domain': 'acbuy.com',       'signup': 'https://acbuy.com/register',                               'coupon': 'Up to $250 in coupons'},
    {'key': 'litbuy',      'name': 'Litbuy',      'build': _b_litbuy,      'color': '#06b6d4', 'domain': 'litbuy.com',      'signup': 'https://www.litbuy.com/register',                          'coupon': 'Up to $150 in coupons'},
    {'key': 'usfans',      'name': 'UsFans',      'build': _b_usfans,      'color': '#dc2626', 'domain': 'usfans.com',      'signup': 'https://www.usfans.com/register',                          'coupon': 'Up to $100 in coupons'},
]


@app.route('/api/signup-agents')
def api_signup_agents():
    """Returns the full agent list with signup URLs + per-agent coupon text."""
    return jsonify({
        'agents': [{
            'key': a['key'],
            'name': a['name'],
            'color': a['color'],
            'domain': a['domain'],
            'signup_url': a.get('signup', ''),
            'coupon': a.get('coupon', 'Welcome credit available'),
            'logo': f'https://www.google.com/s2/favicons?domain={a["domain"]}&sz=64',
        } for a in AGENTS]
    })


def _agents_for_url(seller_url, affcode=''):
    """Build a list of {key, name, url, color, letter, logo} agent options."""
    platform, item_id = _parse_item_url(seller_url)
    out = []
    for a in AGENTS:
        try:
            built = a['build'](seller_url, item_id, platform, affcode)
        except Exception:
            built = None
        if built:
            domain = a.get('domain', '')
            out.append({
                'key': a['key'],
                'name': a['name'],
                'url': built,
                'color': a.get('color', '#444444'),
                'letter': a['name'][0].upper(),
                # Google's favicon service serves brand-icon-quality logos for
                # most domains. Falls back gracefully on the frontend if it
                # fails to load.
                'logo': f'https://www.google.com/s2/favicons?domain={domain}&sz=64' if domain else '',
                'domain': domain,
            })
    return out


def is_admin():
    return session.get('admin_logged_in', False)


def is_admin_api():
    """Either a logged-in admin session OR a valid X-Admin-Token header.

    Master admin calls cross-site APIs with the header version.
    """
    if session.get('admin_logged_in', False):
        return True
    token = request.headers.get('X-Admin-Token') or request.args.get('token')
    if ADMIN_API_TOKEN and token and token == ADMIN_API_TOKEN:
        return True
    return False


@app.context_processor
def inject_config():
    """Make SITE_CONFIG available in all templates as 'site'."""
    return {'site': SITE_CONFIG}


# --- Public Routes ---

@app.route('/')
def home():
    products = get_products()
    categories = get_categories()
    import random
    # Pre-bucket by category in Python so Jinja doesn't run selectattr
    # over 9k+ products N times (was the main rendering bottleneck).
    by_cat = {}
    for p in products:
        by_cat.setdefault(p.get('category', ''), []).append(p)
    # Each section only needs the first 4 — slice now to keep payload small.
    cat_previews = {}
    for cat in categories:
        slug = cat['slug'] if isinstance(cat, dict) else cat.get('slug')
        if slug == 'trending':
            continue
        bucket = by_cat.get(slug, [])
        if bucket:
            cat_previews[slug] = {'name': cat['name'], 'slug': slug,
                                  'count': len(bucket), 'picks': bucket[:4]}
    # Conveyor/hero rows use random picks; sample is O(n) but bounded by 40.
    sample_pool = products if len(products) <= 500 else random.sample(products, 500)
    shuffled = list(sample_pool)
    random.shuffle(shuffled)
    resp = make_response(render_template('home.html',
        products=products,
        featured=products[:8],
        cat_previews=cat_previews,
        conveyor=shuffled[:40],
        hero_products=shuffled[:24],
        categories=categories,
        bundles=_all_bundles_with_picks(products)))
    # Browser caches refreshes for 60s so quick re-loads feel instant.
    # Short window keeps category edits / new products visible quickly.
    resp.headers['Cache-Control'] = 'public, max-age=60'
    return resp


# ============================================================
# Outfit bundles — curated multi-piece looks
# ============================================================
BUNDLE_RECIPES = [
    {'slug': 'off-duty',
     'name': 'Off-duty Sunday',
     'tagline': 'Easy weekend layer — hoodie, soft pants, sneakers, one accessory.',
     'cats': ['hoodies', 'pants', 'accessories', 'shoes']},
    {'slug': 'streetwear',
     'name': 'Streetwear staples',
     'tagline': 'The base streetwear kit — tee, hoodie, pants, jacket, sneakers.',
     'cats': ['shirts', 'hoodies', 'pants', 'jackets', 'shoes']},
    {'slug': 'refined',
     'name': 'Refined casual',
     'tagline': 'Clean lines, neutral palette — shirt, jacket, pants, shoes.',
     'cats': ['shirts', 'jackets', 'pants', 'shoes']},
    {'slug': 'layered',
     'name': 'Layered weekday',
     'tagline': 'Six-piece layered build — jacket, hoodie, shirt, pants, shoes, accessory.',
     'cats': ['jackets', 'hoodies', 'shirts', 'pants', 'shoes', 'accessories']},
]


import re as _re_bundle
# Names to skip when picking for a NON-womens bundle slot — keeps menswear
# bundles from accidentally picking women's leggings, dresses, etc.
_BUNDLE_AVOID_PAT = _re_bundle.compile(
    r'\b(legging|leggings|yoga|alo|sports?\s*bra|bra(?:lette)?|dress|skirt|romper|crop[ -]?top)\b',
    _re_bundle.I,
)
# Also avoid out-of-stock items in bundles
def _bundle_eligible(p, cat):
    if p.get('in_stock') == 0:
        return False
    if cat != 'womens' and _BUNDLE_AVOID_PAT.search(p.get('name', '')):
        return False
    return True


def _bundle_products(recipe, all_products):
    """Pick one product per recipe category, stable across page loads.
    Filters out items inappropriate for the bundle (e.g. women's leggings in
    a non-womens menswear pants slot)."""
    import hashlib
    by_cat = {}
    for p in all_products:
        by_cat.setdefault(p.get('category', ''), []).append(p)
    # Sort each category bucket so picks are deterministic.
    for cat in by_cat:
        by_cat[cat] = sorted(by_cat[cat], key=lambda x: x.get('id', ''))
    picks = []
    used_ids = set()
    for cat in recipe['cats']:
        cands = [p for p in by_cat.get(cat, [])
                 if p.get('id') not in used_ids and _bundle_eligible(p, cat)]
        if not cands:
            # Fall back to the unfiltered bucket if the filter wiped everything
            cands = [p for p in by_cat.get(cat, []) if p.get('id') not in used_ids]
            if not cands:
                continue
        # Hash recipe slug + cat for a stable, distinct pick per bundle slot.
        h = hashlib.md5(f"{recipe['slug']}-{cat}".encode()).hexdigest()
        idx = int(h, 16) % len(cands)
        pick = cands[idx]
        picks.append(pick)
        used_ids.add(pick.get('id'))
    return picks


def _all_bundles_with_picks(all_products):
    out = []
    for r in BUNDLE_RECIPES:
        picks = _bundle_products(r, all_products)
        if len(picks) >= 2:
            total = sum(float(p.get('price_numeric') or 0) for p in picks)
            out.append({**r, 'products': picks, 'count': len(picks), 'total': total})
    return out


@app.route('/bundle/<slug>')
def bundle(slug):
    recipe = next((r for r in BUNDLE_RECIPES if r['slug'] == slug), None)
    if not recipe:
        return redirect(url_for('shop'))
    items = _bundle_products(recipe, get_products())
    total = sum(float(p.get('price_numeric') or 0) for p in items)
    return render_template('bundle.html',
        bundle=recipe,
        products=items,
        total=total,
        categories=get_categories())


@app.route('/shop')
def shop():
    category = request.args.get('category', '')
    sort = request.args.get('sort', 'newest')
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 40

    if q:
        all_products = search_products(q)
        if category == 'trending':
            top_ids = {p['id'] for p in get_top_clicked_products(limit=75)}
            all_products = [p for p in all_products if p['id'] in top_ids]
        elif category:
            all_products = [p for p in all_products if p.get('category') == category]
    elif category == 'trending':
        # Trending = top 75 most-clicked products in the last 30 days (auto-derived
        # from analytics — no manual tagging needed). Falls back to tag-based
        # trending if there aren't enough clicks yet to fill 75 slots.
        all_products = get_top_clicked_products(limit=75)
        if len(all_products) < 75:
            top_ids = {p['id'] for p in all_products}
            tagged = [p for p in get_products() if 'trending' in (p.get('tags') or '').split() and p['id'] not in top_ids]
            all_products.extend(tagged[: 75 - len(all_products)])
    else:
        all_products = get_products(category if category else None)

    if sort == 'price_low':
        all_products.sort(key=lambda p: p.get('price_numeric', 0))
    elif sort == 'price_high':
        all_products.sort(key=lambda p: p.get('price_numeric', 0), reverse=True)

    total = len(all_products)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    products = all_products[(page - 1) * per_page : page * per_page]

    return render_template('shop.html', products=products, categories=get_categories(),
        current_category=category, current_sort=sort, search_query=q,
        page=page, total_pages=total_pages, total=total)


@app.route('/agents')
def agents_page():
    return render_template('agents.html', agents=AGENTS)


@app.route('/link-converter')
def link_converter():
    return render_template('link_converter.html')


@app.route('/tutorial')
def tutorial():
    return render_template('tutorial.html')


@app.route('/spreadsheet')
def spreadsheet():
    products = get_products()
    return render_template('spreadsheet.html', products=products, categories=get_categories(), total=len(products))


@app.route('/privacy')
def privacy():
    return render_template('privacy.html')


@app.route('/terms')
def terms():
    return render_template('terms.html')


@app.route('/robots.txt')
def robots():
    domain = SITE_CONFIG['domain']
    return f"""User-agent: *
Allow: /
Sitemap: https://{domain}/sitemap.xml

User-agent: Googlebot
Allow: /
Crawl-delay: 1

User-agent: ChatGPT-User
Allow: /

User-agent: GPTBot
Allow: /

User-agent: OAI-SearchBot
Allow: /

User-agent: Bingbot
Allow: /
""", 200, {'Content-Type': 'text/plain'}


@app.route('/llms.txt')
def llms_txt():
    s = SITE_CONFIG
    return f"""# {s['name']}

> The #1 {s['agent_name']} finds website. Browse curated products at the best prices. Sign up to {s['agent_name']} for ${s['coupon_amount']} in welcome coupons.

## Pages

- [Home](https://{s['domain']}/): Browse all {s['agent_name']} products
- [Shop](https://{s['domain']}/shop): Search and filter products
- [Link Converter](https://{s['domain']}/link-converter): Convert any agent link to {s['agent_name']}
- [Spreadsheet](https://{s['domain']}/spreadsheet): Full product spreadsheet

## Keywords
{s['agent_name'].lower()}, {s['name'].lower()}, {s['agent_name'].lower()} products, {s['agent_name'].lower()} finds, {s['agent_name'].lower()} coupons, rep finds
""", 200, {'Content-Type': 'text/plain'}


@app.route('/sitemap.xml')
def sitemap():
    domain = SITE_CONFIG['domain']
    pages = ['/', '/shop', '/spreadsheet', '/link-converter', '/privacy', '/terms']
    for cat in get_categories():
        pages.append(f'/shop?category={cat["slug"]}')
    xml = '<?xml version="1.0" encoding="UTF-8"?>\n<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    for page in pages:
        xml += f'  <url><loc>https://{domain}{page}</loc><changefreq>daily</changefreq><priority>{"1.0" if page == "/" else "0.8"}</priority></url>\n'
    xml += '</urlset>'
    return xml, 200, {'Content-Type': 'application/xml'}


@app.route('/go/<product_id>')
def affiliate_redirect(product_id):
    product = get_product(product_id)
    if not product:
        return redirect(url_for('shop'))
    try:
        record_click({
            'product_id': product_id, 'product_name': product.get('name', ''),
            'category': product.get('category', ''), 'element_type': 'affiliate',
            'page': 'redirect', 'referrer': request.referrer or '',
            'user_ip': request.headers.get('X-Forwarded-For', request.remote_addr or ''),
            'user_agent': request.headers.get('User-Agent', ''), 'country': '',
        })
    except Exception:
        pass
    url = product.get('url', '')
    if not url:
        return redirect(url_for('shop'))
    from urllib.parse import urlparse, urlunparse, parse_qsl, urlencode, quote
    agent_domain = (SITE_CONFIG.get('agent_domain') or '').lower()
    parsed = urlparse(url)
    target_host = (parsed.netloc or '').lower()
    affcode = SITE_CONFIG.get('affiliate_code') or ''
    # If the stored URL is already an agent URL (e.g. the spreadsheet shipped
    # kakobuy.com/item/details?url=...), just swap in our own affcode —
    # don't re-wrap it (which used to send KakoBuy a self-referential nested URL).
    if agent_domain and agent_domain in target_host:
        q = parse_qsl(parsed.query, keep_blank_values=True)
        if affcode:
            q = [(k, v) for k, v in q if k != 'affcode']
            q.append(('affcode', affcode))
        agent_url = urlunparse(parsed._replace(query=urlencode(q)))
    else:
        agent_url = f"{SITE_CONFIG['agent_product_url']}?url={quote(url)}"
        if affcode:
            agent_url += f"&affcode={affcode}"
    return redirect(agent_url)


# --- API ---

@app.route('/api/search')
def api_search():
    """Live search for the hero autocomplete. Returns lightweight rows."""
    q = (request.args.get('q') or '').strip()
    limit = min(int(request.args.get('limit', 8) or 8), 20)
    if len(q) < 2:
        return jsonify({'results': []})
    rows = search_products(q)[:limit]
    out = []
    for r in rows:
        out.append({
            'id': r.get('id'),
            'name': r.get('name'),
            'price': r.get('price'),
            'image': r.get('image'),
            'category': r.get('category'),
            'seller': r.get('seller', ''),
        })
    return jsonify({'q': q, 'count': len(out), 'results': out})


@app.route('/api/images/<pid>')
def api_images(pid):
    """Scrape the gallery for a product from its Weidian item page.

    Weidian's HTML embeds a JSON blob with full-resolution
    pcitem<shopId>-<hash>_<W>_<H>.jpg URLs. We pull all of them out and
    return as a list. Cached in-memory per pid.
    """
    p = get_product(pid)
    if not p:
        return jsonify({'images': []})
    cache = api_images._cache
    if pid in cache:
        return jsonify({'images': cache[pid], 'cached': True})
    import re as _re
    import requests as _r
    raw = _unwrap_agent_url(p.get('url', ''))
    images = []
    # Always include the primary image we already have
    primary = p.get('image') or ''
    if primary:
        images.append(primary)
    weidian_id_m = _re.search(r'itemID=(\d+)', raw, _re.I)
    if weidian_id_m:
        wid = weidian_id_m.group(1)
        try:
            r = _r.get(f'https://weidian.com/item.html?itemID={wid}',
                       headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36'},
                       timeout=8)
            html = r.text
            # Extract all pcitem<...>_W_H.{jpg,png,webp} URLs (full-res gallery shots)
            pattern = _re.compile(r'https?://si\.geilicdn\.com/pcitem\d+-[a-f0-9]+_\d+_\d+\.(?:jpg|jpeg|png|webp)', _re.I)
            for m in pattern.findall(html):
                u = m.strip()
                if u not in images:
                    images.append(u)
        except Exception:
            pass
    # Dedupe while preserving order; cap at 12
    seen, dedup = set(), []
    for u in images:
        if u and u not in seen:
            seen.add(u); dedup.append(u)
    dedup = dedup[:12]
    cache[pid] = dedup
    return jsonify({'images': dedup, 'cached': False})
api_images._cache = {}


def _find_ategoat_id(product):
    """Look up the matching ategoat product ID. Cached on disk for 30 days
    since the mapping is effectively permanent per product."""
    pid = product.get('id') or ''
    cache_key = f'aid:{pid}'
    cached = cache_get(cache_key, max_age_seconds=30 * 24 * 3600)
    if cached is not None:
        return cached or None
    import re as _re
    import requests as _r
    url = product.get('url') or ''
    found = None
    m = _re.search(r'itemID(?:%3D|=)(\d+)', url, _re.I)
    if m:
        wid = m.group(1)
        try:
            s = _r.get('https://www.ategoat.com/wp-json/wiligoods/v1/product/list',
                       params={'name': wid, 'limit': 5},
                       headers={'User-Agent': 'Mozilla/5.0', 'Accept': 'application/json'},
                       timeout=6)
            items = ((s.json() or {}).get('data') or {}).get('items') or []
            if items:
                found = items[0].get('id')
        except Exception:
            pass
    if not found:
        name = product.get('name') or ''
        clean = _re.sub(r'^\s*\d+\s*[、,.]\s*', '', name)
        clean = _re.sub(r'\[[^\]]+\]', '', clean).strip()[:60]
        if clean:
            try:
                s = _r.get('https://www.ategoat.com/wp-json/wiligoods/v1/product/list',
                           params={'name': clean, 'limit': 20},
                           headers={'User-Agent': 'Mozilla/5.0', 'Accept': 'application/json'},
                           timeout=6)
                items = ((s.json() or {}).get('data') or {}).get('items') or []
                match = _strict_ategoat_match(product, items)
                if match:
                    found = match.get('id')
            except Exception:
                pass
    cache_set(cache_key, found or '')
    return found


@app.route('/api/variants/<pid>')
def api_variants(pid):
    """Look up the ategoat product page for this item, extract the SKU/variant
    list + gallery images. Cached in-memory.

    Returns {variants: [{title, price, stock, image_url}], images: [...]}
    """
    p = get_product(pid)
    if not p:
        return jsonify({'variants': [], 'images': []})
    cached = cache_get(f'variants:{pid}', max_age_seconds=7 * 24 * 3600)
    if cached is not None:
        return jsonify(cached)
    import re as _re
    import requests as _r
    import json as _json
    result = {'variants': [], 'images': []}
    try:
        aid = _find_ategoat_id(p)
        if aid:
            # Fetch the HTML product page and extract the SKU JSON
            r = _r.get(f'https://www.ategoat.com/product/{aid}',
                       headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36'},
                       timeout=10)
            if r.status_code == 200:
                html = r.text
                m = _re.search(r'currentProductSkuItems\s*=\s*(\[.*?\]);', html, _re.S)
                if m:
                    try:
                        skus = _json.loads(m.group(1))
                        # Dedupe to one SKU per unique image (i.e. per colour/
                        # style) — strip out per-size duplicates so the user
                        # sees one button per actual variant, not a button per
                        # size that all look identical.
                        seen_imgs = set()
                        deduped = []
                        for s in skus:
                            img = s.get('image_url') or ''
                            if not img or img in seen_imgs:
                                continue
                            seen_imgs.add(img)
                            # Strip the trailing ";SIZE" off the title so the
                            # label reads as a colour/style code only.
                            t = str(s.get('title') or '')
                            t_clean = t.split(';')[0].strip() if ';' in t else t
                            deduped.append({
                                'title': t_clean or t,
                                'price': s.get('price'),
                                'stock': s.get('stock', 0),
                                'image': img,
                            })
                            if len(deduped) >= 30:
                                break
                        result['variants'] = deduped
                    except Exception:
                        pass
                # Also pull any product images out of the HTML — useful when
                # the Weidian scrape failed for this product.
                img_re = _re.compile(r'https?://(?:si|sd)\.geilicdn\.com/[a-zA-Z0-9_/.\-]+\.(?:jpg|jpeg|png|webp)', _re.I)
                imgs_raw = img_re.findall(html)
                imgs = []
                seen = set()
                for u in imgs_raw:
                    if any(skip in u for skip in ('hz_img_', 'icon-', '/avatar', 'login_', 'wd_logo', 'common-')):
                        continue
                    if u in seen: continue
                    seen.add(u); imgs.append(u)
                    if len(imgs) >= 12: break
                result['images'] = imgs
    except Exception:
        pass
    cache_set(f'variants:{pid}', result)
    return jsonify(result)


def _strict_ategoat_match(product, candidates):
    """Pick an ategoat product that is *actually* the same item.

    Strict criteria (returns None unless one is met):
      1) Our primary image URL appears in the candidate's image (strong signal —
         same source photo means same listing).
      2) The candidate URL (anywhere we can find one) contains our Weidian
         itemID.
      3) Token overlap >= 60% of OUR name tokens AND the candidate title
         starts with the same first 1-2 tokens (so 'Moncler Down Jacket'
         doesn't blindly match the first 'Moncler' result).
    """
    import re as _re
    if not candidates:
        return None
    our_img = (product.get('image') or '').split('?')[0]
    our_name = (product.get('name') or '').strip().lower()
    our_tokens = [t for t in _re.split(r'[^a-z0-9]+', our_name) if len(t) > 1]
    our_url = product.get('url') or ''
    wid_m = _re.search(r'itemID(?:%3D|=)(\d+)', our_url, _re.I)
    our_wid = wid_m.group(1) if wid_m else ''

    # 1) Image-URL match
    if our_img:
        for it in candidates:
            cand_img = (it.get('image') or '').split('?')[0]
            if cand_img and (our_img == cand_img or our_img in cand_img or cand_img in our_img):
                return it

    # 2) Weidian itemID present in any candidate URL field
    if our_wid:
        for it in candidates:
            for k in ('source_url', 'url', 'original_url', 'goods_url', 'title'):
                v = str(it.get(k) or '')
                if our_wid in v:
                    return it

    # 2) Token-overlap match — must clear the bar
    if our_tokens:
        our_set = set(our_tokens)
        first_two = ' '.join(our_tokens[:2]).lower()
        best = None
        best_score = 0
        for it in candidates:
            t = (it.get('title') or '').lower()
            ct = [x for x in _re.split(r'[^a-z0-9]+', t) if len(x) > 1]
            if not ct:
                continue
            overlap = len(our_set & set(ct))
            ratio = overlap / max(1, len(our_set))
            if ratio < 0.6:
                continue
            if not t.startswith(our_tokens[0]) and first_two not in t:
                continue
            if overlap > best_score:
                best_score = overlap
                best = it
        if best:
            return best
    return None


@app.route('/api/qc/<pid>')
def api_qc(pid):
    """QC-photo fetch from ategoat.com — uses itemID-first matcher so we never
    show QC photos that belong to a different listing."""
    p = get_product(pid)
    if not p:
        return jsonify({'photos': []})
    cached = cache_get(f'qc:{pid}', max_age_seconds=7 * 24 * 3600)
    if cached is not None:
        return jsonify({'photos': cached})
    import requests as _r
    photos = []
    try:
        aid = _find_ategoat_id(p)
        if aid:
            q = _r.get('https://www.ategoat.com/wp-json/wiligoods/v1/qc/list',
                       params={'product_id': aid, 'limit': 30},
                       headers={'User-Agent': 'Mozilla/5.0', 'Accept': 'application/json'},
                       timeout=6)
            qitems = ((q.json() or {}).get('data') or {}).get('items') or []
            photos = [it.get('image') or it.get('url') for it in qitems if it.get('image') or it.get('url')]
    except Exception:
        photos = []
    cache_set(f'qc:{pid}', photos)
    return jsonify({'photos': photos})


@app.route('/api/agents/<pid>')
def api_agents(pid):
    """Return per-agent buy URLs for a single product (used by card picker)."""
    p = get_product(pid)
    if not p:
        return jsonify({'error': 'not found'}), 404
    raw = _unwrap_agent_url(p.get('url', ''))
    affcode = SITE_CONFIG.get('affiliate_code', '')
    agents = _agents_for_url(raw, affcode)
    platform, item_id = _parse_item_url(raw)
    return jsonify({
        'product_id': pid,
        'product_name': p.get('name', ''),
        'platform': platform,
        'item_id': item_id,
        'default_agent': SITE_CONFIG.get('agent_name', ''),
        'agents': agents,
    })


@app.route('/api/agents/from-url')
def api_agents_from_url():
    """Build agent URLs from any pasted seller URL. ?url=<raw_or_wrapper>"""
    url = (request.args.get('url') or '').strip()
    if not url:
        return jsonify({'error': 'Not a valid URL', 'agents': [], 'platform': None, 'item_id': None})
    # Reject obvious non-URLs early so the converter can show a clear message.
    if not (url.startswith('http://') or url.startswith('https://')) and not _looks_like_seller_url(url):
        return jsonify({'error': 'Not a valid URL', 'agents': [], 'platform': None, 'item_id': None})
    raw = _unwrap_agent_url(url)
    platform, item_id = _parse_item_url(raw)
    if not platform or not item_id:
        return jsonify({'error': 'Not a valid URL', 'agents': [], 'platform': None, 'item_id': None, 'raw_url': raw})
    affcode = SITE_CONFIG.get('affiliate_code', '')
    agents = _agents_for_url(raw, affcode)
    return jsonify({
        'platform': platform,
        'item_id': item_id,
        'raw_url': raw,
        'agents': agents,
    })


def _looks_like_seller_url(s: str) -> bool:
    """Loose check for a paste that might be a partial URL (e.g. user dropped
    the scheme). Accept anything that mentions a known seller/agent domain."""
    s = s.lower()
    for d in ('weidian.', 'taobao.', '1688.', 'tmall.', 'kakobuy.', 'oopbuy.',
              'sugargoo.', 'joyagoo.', 'allchinabuy.', 'hubbuycn.', 'mulebuy.',
              'acbuy.', 'litbuy.', 'hipobuy.', 'hoobuy.', 'usfans.', 'cnfans.'):
        if d in s:
            return True
    return False


def _related_products(p, limit=8):
    """'You might also like' — same brand, then same category, best-seller nudge."""
    if not p:
        return []
    try:
        from tag_utils import BRANDS
    except Exception:
        BRANDS = {}
    hay = (p.get('name', '') + ' ' + p.get('tags', '')).lower()
    hay_tokens = set(hay.split())
    brand_tokens = set()
    for key, aliases in BRANDS.items():
        toks = aliases if isinstance(aliases, list) else [aliases]
        if key in hay or any(t in hay_tokens for t in toks):
            brand_tokens.update(toks)
    cat = p.get('category', '')
    spid = p.get('id')
    def has_brand(x):
        xt = (x.get('name', '') + ' ' + x.get('tags', '')).lower()
        return any(t in xt for t in brand_tokens)
    def score(x):
        sc = 0.0
        if brand_tokens and has_brand(x): sc += 10
        if x.get('category') == cat: sc += 3
        sc += min(int(x.get('sales') or 0), 2000) / 2000.0
        return sc
    pool = [x for x in get_products()
            if x.get('id') != spid
            and (x.get('category') == cat or (brand_tokens and has_brand(x)))]
    return sorted(pool, key=score, reverse=True)[:limit]


@app.route('/product/<pid>')
def product_page(pid):
    """Standalone on-site product page (same design as the detail modal) + related."""
    p = get_product(pid)
    if not p:
        return redirect(url_for('shop'))
    return render_template('product.html', product=p,
                           related=_related_products(p, 8),
                           categories=get_categories())


@app.route('/api/product/<pid>')
def api_product(pid):
    """Return a product with its listing variants and (stubbed) QC photos.
    Used by the product detail modal."""
    p = get_product(pid)
    if not p:
        return jsonify({'error': 'not found'}), 404
    variants = get_listing_variants(p)
    # If no real variants found (e.g. demo data has no listing_id), still
    # return the product itself as the sole variant.
    if not variants:
        variants = [p]
    # QC photos placeholder — will be wired to Kakobuy QC API later.
    qc_photos = []
    if p.get('qc_photos'):
        try:
            qc_photos = json.loads(p['qc_photos'])
        except (ValueError, TypeError):
            qc_photos = []
    # Enrich missing fields from what we can derive:
    #  - seller: a real Weidian shop ID when we can dig it out of the image
    #    URL (geilicdn embeds it as pcitem<shopId>-...). The catch-all is
    #    the marketplace name (Weidian / Taobao / 1688) when we can't pin
    #    a specific shop.
    #  - batch:  pull out a [XX BATCH] tag baked into many product names
    #  - weight: per-category estimate when no shipping weight is set
    import re
    raw_url = _unwrap_agent_url(p.get('url', ''))
    img_url = p.get('image', '') or ''
    # Try to extract a Weidian shop ID. Geilicdn images encode it as the
    # number after one of several known prefixes (pcitem / open / item /
    # pccover), e.g. pcitem1809160355-..., open1807578469-1234478995-...
    shop_id = None
    sm = re.search(r'geilicdn\.com/(?:pcitem|open|pccover|item|si)(\d{6,})', img_url)
    if sm:
        shop_id = sm.group(1)
    # The shop subdomain pattern shopNNN.v.weidian.com is more reliable when
    # the saved URL exposes it.
    if not shop_id:
        sm2 = re.search(r'shop(\d+)\.v\.weidian', raw_url)
        if sm2:
            shop_id = sm2.group(1)

    derived_seller = p.get('seller') or ''
    seller_url = ''
    marketplace = ''
    if 'weidian.com' in raw_url.lower(): marketplace = 'Weidian'
    elif 'taobao.com' in raw_url.lower() or 'tmall.com' in raw_url.lower(): marketplace = 'Taobao'
    elif '1688.com' in raw_url.lower(): marketplace = '1688'

    if not derived_seller:
        if shop_id and marketplace == 'Weidian':
            derived_seller = f'Weidian shop #{shop_id}'
            seller_url = f'https://shop{shop_id}.v.weidian.com/'
        elif marketplace:
            derived_seller = f'{marketplace} marketplace'
    derived_batch = p.get('batch') or ''
    if not derived_batch:
        m = re.search(r'\[([A-Z]{1,4})\s*BATCH\]', p.get('name', ''), re.I)
        if m:
            derived_batch = m.group(1).upper() + ' Batch'

    # Per-category shipping-weight estimates (grams, packed). Pulled from typical
    # repping community benchmarks — close enough for shipping-cost ballparks.
    WEIGHT_ESTIMATES = {
        'shoes':       '900g',
        'shirts':      '230g',
        'hoodies':     '650g',
        'pants':       '550g',
        'jackets':     '900g',
        'accessories': '180g',
        'bags':        '700g',
        'tech':        '400g',
        'womens':      '320g',
        'trending':    '450g',
    }
    derived_weight = (p.get('weight') or '').strip()
    weight_is_estimate = False
    if not derived_weight:
        derived_weight = WEIGHT_ESTIMATES.get(p.get('category') or '', '450g')
        weight_is_estimate = True

    derived_sales = p.get('sales') or 0
    return jsonify({
        'id': p['id'],
        'name': p['name'],
        'price': p['price'],
        'price_numeric': p.get('price_numeric', 0),
        'image': p['image'],
        'category': p.get('category', ''),
        'seller': derived_seller,
        'seller_url': seller_url,
        'marketplace': marketplace,
        'batch': derived_batch,
        'retail_price': p.get('retail_price', ''),
        'tags': p.get('tags', ''),
        'weight': derived_weight,
        'weight_estimate': weight_is_estimate,
        'quality': p.get('quality', ''),
        'sales': derived_sales,
        'go_url': f'/go/{pid}',
        'raw_url': raw_url,
        'agent_url': p.get('url', ''),
        'images': p.get('images') or [],
        'variants': [
            {
                'id': v['id'],
                'name': v['name'],
                'image': v['image'],
                'price': v['price'],
                'price_numeric': v.get('price_numeric', 0),
            } for v in variants
        ],
        'qc_photos': qc_photos,
    })


@app.route('/api/products')
def api_products():
    limit = request.args.get('limit', 50, type=int)
    products = get_products()
    import random
    sampled = random.sample(products, min(limit, len(products)))
    return jsonify([{'id': p['id'], 'name': p['name'], 'image': p['image'], 'price': p['price'], 'category': p['category']} for p in sampled])


@app.route('/api/click', methods=['POST'])
def api_click():
    data = request.get_json(silent=True) or {}
    data['user_ip'] = request.headers.get('X-Forwarded-For', request.remote_addr or '')
    data['user_agent'] = request.headers.get('User-Agent', '')
    try:
        record_click(data)
    except Exception:
        pass
    return jsonify({'ok': True})


@app.route('/api/track-pageview', methods=['POST'])
def api_pageview():
    data = request.get_json(silent=True) or {}
    data['element_type'] = 'pageview'
    data['user_ip'] = request.headers.get('X-Forwarded-For', request.remote_addr or '')
    data['user_agent'] = request.headers.get('User-Agent', '')
    try:
        record_click(data)
    except Exception:
        pass
    return jsonify({'ok': True})


# --- Admin ---

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if request.form.get('password') == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            session.permanent = True   # honor the 30-day session lifetime
            return redirect(url_for('admin_dashboard'))
        return render_template('admin_login.html', error='Wrong password')
    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('home'))


@app.route('/admin')
def admin_dashboard():
    if not is_admin():
        return redirect(url_for('admin_login'))
    stats = get_analytics(30)
    products = get_products()
    categories = get_categories()
    featured_count = sum(1 for p in products if p.get('featured'))
    cats_with_counts = []
    for c in categories:
        n = sum(1 for p in products if p.get('category') == c.get('slug'))
        cats_with_counts.append({**c, 'count': n})
    recent_products = sorted(products, key=lambda p: p.get('updated_at') or '', reverse=True)[:8]
    return render_template(
        'admin_dashboard.html',
        stats=stats,
        products=products,
        categories=cats_with_counts,
        featured_count=featured_count,
        recent_products=recent_products,
        total_products=len(products),
    )


@app.route('/admin/products/add', methods=['POST'])
def admin_add_product():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    if not data.get('name'):
        return jsonify({'error': 'Name required'}), 400
    data['id'] = data.get('id', f"p{secrets.token_hex(4)}")
    try:
        from tag_utils import generate_tags, auto_category
    except ImportError:
        from .tag_utils import generate_tags, auto_category
    # Auto-categorize if not set: name 'Stussy Hoodie' -> 'hoodies' etc.
    if not (data.get('category') or '').strip():
        data['category'] = auto_category(data['name'], fallback='')
    if not data.get('tags'):
        data['tags'] = generate_tags(data['name'], data.get('category', ''))
    add_product(data)
    return jsonify({'ok': True, 'id': data['id'], 'category': data.get('category', '')})


@app.route('/admin/products/update/<pid>', methods=['POST'])
def admin_update_product(pid):
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({'error': 'No data'}), 400
    if 'price' in data:
        try:
            data['price_numeric'] = float(data['price'] or 0)
        except ValueError:
            data['price_numeric'] = 0
    update_product(pid, data)
    return jsonify({'ok': True})


@app.route('/admin/products/bulk', methods=['POST'])
def admin_bulk():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    products = data.get('products', [])
    if not products:
        return jsonify({'error': 'No products'}), 400
    try:
        from tag_utils import generate_tags, auto_category
    except ImportError:
        from .tag_utils import generate_tags, auto_category
    for p in products:
        if not p.get('id'):
            p['id'] = f"p{secrets.token_hex(4)}"
        if not (p.get('category') or '').strip():
            p['category'] = auto_category(p.get('name', ''), fallback='')
        if not p.get('tags'):
            p['tags'] = generate_tags(p.get('name', ''), p.get('category', ''))
    add_products_bulk(products)
    return jsonify({'ok': True, 'count': len(products)})


@app.route('/admin/products/delete/<pid>', methods=['DELETE'])
def admin_delete(pid):
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    delete_product(pid)
    return jsonify({'ok': True})


@app.route('/admin/products/delete-batch', methods=['POST'])
def admin_delete_batch():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    for pid in ids:
        delete_product(pid)
    return jsonify({'ok': True, 'count': len(ids)})


@app.route('/admin/products/feature', methods=['POST'])
def admin_feature():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    featured = bool(data.get('featured', True))
    n = set_featured(ids, featured)
    return jsonify({'ok': True, 'count': n, 'featured': featured})


@app.route('/admin/products/move-category', methods=['POST'])
def admin_move_category():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    category = (data.get('category') or '').strip()
    if not category:
        return jsonify({'error': 'Category required'}), 400
    n = move_category(ids, category)
    return jsonify({'ok': True, 'count': n, 'category': category})


@app.route('/admin/products/reorder', methods=['POST'])
def admin_reorder():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    n = reorder_products(ids)
    return jsonify({'ok': True, 'count': n})


@app.route('/admin/products')
def admin_products():
    if not is_admin():
        return redirect(url_for('admin_login'))
    # Silent auto-categorize: every product without a category whose name has
    # a clear hint gets moved into its proper bucket. Idempotent + fast.
    try:
        try:
            from tag_utils import auto_category
        except ImportError:
            from .tag_utils import auto_category
        products = get_products()
        for p in products:
            if (p.get('category') or '').strip(): continue
            guess = auto_category(p.get('name', ''), fallback='')
            if guess:
                update_product(p['id'], {'category': guess})
    except Exception:
        pass
    # Render only the first 200 rows server-side; the rest stream in via
    # /admin/products/rows on demand. This cut admin HTML from ~35MB to
    # ~800KB for the 9k-product catalogue.
    all_products = get_products()
    INITIAL = 200
    return render_template('admin_products.html',
        products=all_products[:INITIAL],
        total_products=len(all_products),
        initial_count=min(INITIAL, len(all_products)),
        categories=get_categories())


@app.route('/admin/products/rows')
def admin_products_rows():
    """Lazy-load additional rows of the admin table beyond the initial 200."""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    offset = max(0, int(request.args.get('offset', 200)))
    limit = min(2000, max(50, int(request.args.get('limit', 1000))))
    all_products = get_products()
    chunk = all_products[offset:offset+limit]
    return render_template('admin_product_rows.html',
        products=chunk, categories=get_categories())


@app.route('/admin/products/uncategorized')
def admin_list_uncategorized():
    """Return all products with an empty category. Used by the Uncategorized tab."""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    products = get_products()
    items = [
        {'id': p['id'], 'name': p.get('name',''), 'image': p.get('image',''),
         'tags': p.get('tags','')}
        for p in products if not (p.get('category') or '').strip()
    ]
    return jsonify({'count': len(items), 'products': items})


@app.route('/admin/analytics')
def admin_analytics():
    if not is_admin():
        return redirect(url_for('admin_login'))
    from datetime import datetime, timedelta

    range_key = request.args.get('range', '30d')
    now = datetime.now()
    since = None
    until = now

    if range_key == 'today':
        since = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif range_key == 'hour':
        since = now - timedelta(hours=1)
    elif range_key == '24h':
        since = now - timedelta(hours=24)
    elif range_key == '7d':
        since = now - timedelta(days=7)
    elif range_key == '30d':
        since = now - timedelta(days=30)
    elif range_key == '90d':
        since = now - timedelta(days=90)
    elif range_key == 'mtd':  # Month-to-date: 1st of this month → now
        since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif range_key == 'ytd':  # Year-to-date: Jan 1 → now
        since = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    elif range_key == '1y':
        since = now - timedelta(days=365)
    elif range_key == 'custom':
        f = request.args.get('from')
        t = request.args.get('to')
        if f:
            try: since = datetime.fromisoformat(f)
            except ValueError: pass
        if t:
            try:
                # Treat "to" as end-of-day so 2026-05-11 includes all of that day
                until_dt = datetime.fromisoformat(t)
                until = until_dt.replace(hour=23, minute=59, second=59, microsecond=999999) if until_dt.hour == 0 and until_dt.minute == 0 else until_dt
            except ValueError: pass

    if since is None:
        since = now - timedelta(days=30)
        range_key = '30d'

    stats = get_analytics(since=since.isoformat(), until=until.isoformat())
    return render_template(
        'admin_analytics.html',
        stats=stats,
        range_key=range_key,
        range_since=since.strftime('%Y-%m-%d'),
        range_until=until.strftime('%Y-%m-%d'),
        # Hours-in-range for "Daily Avg" math
        span_seconds=(until - since).total_seconds(),
    )


@app.route('/admin/analytics/export')
def admin_analytics_export():
    """Export one of the analytics breakdowns as CSV."""
    if not is_admin():
        return redirect(url_for('admin_login'))
    from datetime import datetime, timedelta
    from io import StringIO
    import csv

    kind = request.args.get('kind', 'products')
    range_key = request.args.get('range', '30d')
    now = datetime.now()
    until = now
    if range_key == 'today':   since = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif range_key == 'hour':  since = now - timedelta(hours=1)
    elif range_key == '24h':   since = now - timedelta(hours=24)
    elif range_key == '7d':    since = now - timedelta(days=7)
    elif range_key == '30d':   since = now - timedelta(days=30)
    elif range_key == '90d':   since = now - timedelta(days=90)
    elif range_key == 'mtd':   since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif range_key == 'ytd':   since = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    elif range_key == '1y':    since = now - timedelta(days=365)
    elif range_key == 'custom':
        try: since = datetime.fromisoformat(request.args.get('from') or '')
        except (ValueError, TypeError): since = now - timedelta(days=30)
        try: until = datetime.fromisoformat(request.args.get('to') or '').replace(hour=23, minute=59, second=59)
        except (ValueError, TypeError): pass
    else:                      since = now - timedelta(days=30)

    stats = get_analytics(since=since.isoformat(), until=until.isoformat())
    fmt = (request.args.get('format') or 'csv').lower()

    # Build rows (kind → (headers, list-of-rows))
    if kind == 'products':
        headers = ['Rank', 'Product', 'Clicks']
        rows = [[i, p['product_name'], p['clicks']] for i, p in enumerate(stats['top_products'], 1)]
    elif kind == 'categories':
        headers = ['Rank', 'Category', 'Clicks']
        rows = [[i, c['category'], c['clicks']] for i, c in enumerate(stats['top_categories'], 1)]
    elif kind == 'pages':
        headers = ['Rank', 'Page', 'Views']
        rows = [[i, p['page'], p['views']] for i, p in enumerate(stats['top_pages'], 1)]
    elif kind == 'types':
        headers = ['Rank', 'Element Type', 'Clicks']
        rows = [[i, e['element_type'] or 'unknown', e['clicks']] for i, e in enumerate(stats['element_types'], 1)]
    elif kind == 'daily':
        headers = ['Date', 'Clicks', 'Visitors']
        rows = [[d['day'], d['clicks'], d['visitors']] for d in stats['daily']]
    else:
        return jsonify({'error': 'unknown kind'}), 400

    base_name = f'{SITE_CONFIG.get("name","kai").lower().replace(" ","-")}-{kind}-{range_key}-{since.strftime("%Y%m%d")}'

    from flask import make_response
    if fmt in ('xlsx', 'excel'):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return jsonify({'error': 'openpyxl not installed on server — pip install openpyxl'}), 500
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = kind.title()
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='06B6D4')
            cell.alignment = Alignment(horizontal='left')
        for r in rows:
            ws.append(r)
        # Auto-size columns
        for col_idx, h in enumerate(headers, 1):
            longest = max([len(str(h))] + [len(str(r[col_idx-1])) for r in rows], default=10)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(longest + 4, 60)
        ws.freeze_panes = 'A2'
        buf = BytesIO()
        wb.save(buf)
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="{base_name}.xlsx"'
        return resp

    # CSV (default)
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = f'attachment; filename="{base_name}.csv"'
    return resp


@app.route('/admin/categories/add', methods=['POST'])
def admin_add_category():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    if not data.get('slug') or not data.get('name'):
        return jsonify({'error': 'Slug and name required'}), 400
    add_category(data['slug'], data['name'], data.get('icon', ''), data.get('description', ''), data.get('sort_order', 0))
    return jsonify({'ok': True})


@app.route('/admin/categories')
def admin_list_categories():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    cats = get_categories()
    for c in cats:
        c['product_count'] = count_products_in_category(c['slug'])
    return jsonify({'categories': cats})


@app.route('/admin/categories/update/<slug>', methods=['POST'])
def admin_update_category(slug):
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    new_slug = update_category(slug, data)
    return jsonify({'ok': True, 'slug': new_slug})


@app.route('/admin/categories/delete/<slug>', methods=['POST'])
def admin_delete_category(slug):
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    reassign_to = (data.get('reassign_to') or '').strip()
    delete_category(slug, reassign_to)
    return jsonify({'ok': True})


@app.route('/admin/products/set-category', methods=['POST'])
def admin_set_one_category():
    """Quick-set a single product's category from the table dropdown."""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    pid = (data.get('id') or '').strip()
    category = (data.get('category') or '').strip()
    if not pid:
        return jsonify({'error': 'Product id required'}), 400
    update_product(pid, {'category': category})
    return jsonify({'ok': True, 'id': pid, 'category': category})


@app.route('/admin/products/stock', methods=['POST'])
def admin_set_stock():
    """Toggle in_stock on one or many products. Body: {ids:[pid,...], in_stock:bool}."""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    ids = data.get('ids') or []
    if isinstance(ids, str):
        ids = [ids]
    in_stock = bool(data.get('in_stock', True))
    if not ids:
        return jsonify({'error': 'No product ids'}), 400
    n = set_in_stock(ids, in_stock)
    return jsonify({'ok': True, 'count': n, 'in_stock': in_stock})


@app.route('/admin/products/trending', methods=['POST'])
def admin_trending():
    """Add or remove the 'trending' tag for one or more products. Products with
    the tag show up under /shop?category=trending in addition to their real
    category."""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    ids = data.get('ids') or []
    if isinstance(ids, str):
        ids = [ids]
    trending = bool(data.get('trending', True))
    if not ids:
        return jsonify({'error': 'No product ids'}), 400
    n = 0
    for pid in ids:
        p = get_product(pid)
        if not p:
            continue
        tags = (p.get('tags') or '').split()
        had = 'trending' in tags
        if trending and not had:
            tags.append('trending')
        elif not trending and had:
            tags = [t for t in tags if t != 'trending']
        else:
            continue
        update_product(pid, {'tags': ' '.join(tags)})
        n += 1
    return jsonify({'ok': True, 'count': n, 'trending': trending})


@app.route('/admin/backup', methods=['POST'])
def admin_backup():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify({'ok': True, 'path': backup_database()})


@app.route('/admin/scrape', methods=['POST'])
def admin_scrape():
    """Scrape a single Weidian/Taobao/1688 listing → returns {products, listing_name,
    total_variants, platform, item_id}. On failure returns {error: '<human msg>'}."""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    url = (data.get('url') or '').strip()
    category = (data.get('category') or '').strip()
    if not url:
        return jsonify({'error': 'No URL provided. Paste a Weidian, Taobao, or 1688 product link.'}), 400
    # Quick sanity: must look like a URL
    if not (url.startswith('http://') or url.startswith('https://')):
        return jsonify({'error': 'Not a valid URL. Must start with http:// or https://'}), 400
    try:
        try:
            from .scraper import scrape_listing, detect_platform
        except ImportError:
            from scraper import scrape_listing, detect_platform
        platform, item_id = detect_platform(url)
        if not platform:
            return jsonify({'error': 'Could not parse this link. Supported: Weidian, Taobao, 1688. (Looked for itemID=, id=, /offer/N.html.)'}), 400
        result = scrape_listing(url, category=category, affiliate_code=SITE_CONFIG.get('affiliate_code', ''))
        if not result or not isinstance(result, dict):
            return jsonify({'error': f'Scraper returned no data for {platform} item {item_id}. The listing may be private, expired, or geo-blocked.'}), 502
        if result.get('error'):
            return jsonify({'error': result['error']}), 502
        products = result.get('products') or []
        if not products:
            return jsonify({'error': f'No products / variants extracted from {platform} item {item_id}. The seller may have deleted the listing or blocked scraping.'}), 502
        return jsonify(result)
    except ImportError as e:
        return jsonify({'error': f'Scraper module not available on the server: {e}'}), 500
    except Exception as e:
        return jsonify({'error': f'Scrape failed: {type(e).__name__}: {str(e)[:200]}'}), 500


@app.route('/admin/scrape/import', methods=['POST'])
def admin_scrape_import():
    """Commit selected scraped products into the catalog."""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    products = data.get('products') or []
    if not products:
        return jsonify({'error': 'Nothing to import'}), 400
    try:
        from tag_utils import generate_tags, auto_category
    except ImportError:
        from .tag_utils import generate_tags, auto_category
    for p in products:
        if not p.get('id'):
            p['id'] = f"p{secrets.token_hex(4)}"
        if not (p.get('category') or '').strip():
            p['category'] = auto_category(p.get('name', ''), fallback='')
        if not p.get('tags'):
            p['tags'] = generate_tags(p.get('name', ''), p.get('category', ''))
    add_products_bulk(products)
    return jsonify({'ok': True, 'count': len(products)})


@app.route('/admin/products/auto-categorize', methods=['POST'])
def admin_auto_categorize_all():
    """Apply auto_category(name) to every product currently without a category.
    Returns how many got updated. Idempotent."""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        from tag_utils import auto_category
    except ImportError:
        from .tag_utils import auto_category
    products = get_products()
    updated = 0
    for p in products:
        if (p.get('category') or '').strip():
            continue
        guess = auto_category(p.get('name', ''), fallback='')
        if guess:
            update_product(p['id'], {'category': guess})
            updated += 1
    return jsonify({'ok': True, 'updated': updated, 'scanned': len(products)})


# ===========================================================================
# Cross-site API (used by the master admin) — token-auth alternative.
# All routes accept either an admin session OR an X-Admin-Token header.
# ===========================================================================

@app.route('/admin/api/ping')
def api_ping():
    """Simple liveness check that also validates the token. Public side."""
    has_token = bool(ADMIN_API_TOKEN)
    token_ok = is_admin_api()
    return jsonify({
        'ok': True,
        'site': SITE_CONFIG.get('name'),
        'agent': SITE_CONFIG.get('agent_name'),
        'token_required': has_token,
        'token_valid': token_ok,
    })


@app.route('/admin/api/stats')
def api_stats():
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    from datetime import datetime, timedelta
    days = request.args.get('days', 30, type=int)
    since = (datetime.now() - timedelta(days=days)).isoformat()
    stats = get_analytics(days=days)
    products = get_products()
    return jsonify({
        'site': SITE_CONFIG.get('name'),
        'agent': SITE_CONFIG.get('agent_name'),
        'total_products': len(products),
        'featured_count': sum(1 for p in products if p.get('featured')),
        'categories': len(get_categories()),
        'total_clicks': stats['total_clicks'],
        'unique_visitors': stats['unique_visitors'],
        'signup_clicks': stats['signup_clicks'],
        'top_products': stats['top_products'],
        'top_categories': stats['top_categories'],
        'daily': stats['daily'],
        'days': days,
        'since': since,
    })


@app.route('/admin/api/products')
def api_admin_products():
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify({
        'site': SITE_CONFIG.get('name'),
        'count': None,
        'products': get_products(),
    })


@app.route('/admin/api/products', methods=['POST'])
def api_admin_add_product():
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    if not data.get('name'):
        return jsonify({'error': 'Name required'}), 400
    if not data.get('id'):
        data['id'] = f"p{secrets.token_hex(4)}"
    if not data.get('tags'):
        try:
            try:
                from .tag_utils import generate_tags
            except ImportError:
                from tag_utils import generate_tags
            data['tags'] = generate_tags(data['name'], data.get('category', ''))
        except ImportError:
            pass
    add_product(data)
    return jsonify({'ok': True, 'id': data['id']})


@app.route('/admin/api/products/<pid>', methods=['PUT', 'PATCH'])
def api_admin_update_product(pid):
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    if 'price' in data:
        try: data['price_numeric'] = float(data['price'] or 0)
        except ValueError: data['price_numeric'] = 0
    update_product(pid, data)
    return jsonify({'ok': True})


@app.route('/admin/api/products/<pid>', methods=['DELETE'])
def api_admin_delete_product(pid):
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    delete_product(pid)
    return jsonify({'ok': True})


@app.route('/admin/api/config')
def api_admin_config():
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    # Don't leak secrets — strip token-ish keys
    safe = {k: v for k, v in SITE_CONFIG.items() if 'token' not in k.lower() and 'secret' not in k.lower()}
    return jsonify(safe)


@app.route('/admin/backup/download')
def admin_download_backup():
    if not is_admin():
        return redirect(url_for('admin_login'))
    return send_file(backup_database(), as_attachment=True)


# ===========================================================================
# Custom image uploads — stored on the persistent disk so they survive deploys
# ===========================================================================
UPLOADS_DIR = os.path.join(DATA_DIR, 'uploads')
os.makedirs(UPLOADS_DIR, exist_ok=True)
ALLOWED_IMG_EXT = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}


@app.route('/uploads/<path:fname>')
def serve_upload(fname):
    """Serve uploaded images from the persistent disk."""
    # Path-traversal guard
    safe = os.path.basename(fname)
    full = os.path.join(UPLOADS_DIR, safe)
    if not os.path.exists(full):
        return ('', 404)
    return send_file(full)


@app.route('/admin/products/upload-image/<pid>', methods=['POST'])
def admin_upload_image(pid):
    """Accept a multipart image upload, save to disk, set as product's primary image."""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f or not f.filename:
        return jsonify({'error': 'Empty filename'}), 400
    import os as _os
    ext = _os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_IMG_EXT:
        return jsonify({'error': f'Unsupported image type ({ext}). Use jpg/png/webp/gif.'}), 400
    # Size cap: 8MB
    f.seek(0, 2); size = f.tell(); f.seek(0)
    if size > 8 * 1024 * 1024:
        return jsonify({'error': 'File too large (max 8MB)'}), 400
    # Save with safe name
    safe_pid = ''.join(c for c in pid if c.isalnum())[:16] or 'p'
    fname = f'{safe_pid}-{secrets.token_hex(4)}{ext}'
    full = _os.path.join(UPLOADS_DIR, fname)
    f.save(full)
    new_url = f'/uploads/{fname}'
    update_product(pid, {'image': new_url})
    return jsonify({'ok': True, 'image': new_url, 'pid': pid})


@app.errorhandler(404)
def not_found(e):
    return redirect(url_for('home'))


if __name__ == '__main__':
    os.makedirs(DATA_DIR, exist_ok=True)
    app.run(debug=True, port=int(os.environ.get('PORT', 5010)))
