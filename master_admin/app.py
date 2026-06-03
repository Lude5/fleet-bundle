"""Master Admin — multi-tenant control plane for all client fashion-finds sites.

Inspired by ategoat's tenant dashboard. Runs locally by default on :5050,
manages sites configured in sites.json. Calls each client site's admin API
via X-Admin-Token header (once Phase 3 ships) — for now reads sites.json
state and links out to each client's /admin.
"""
import os
import json
import secrets
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path

from flask import (Flask, render_template, request, jsonify, redirect,
                   session, url_for, send_from_directory)

ROOT = Path(__file__).resolve().parent       # _master_admin/
CLIENTS = ROOT.parent                         # clients/
SITES_FILE = ROOT / 'sites.json'


def _sibling(name):
    """Import a sibling module (ga_client, scrape_product) whether app.py runs
    directly (cwd=master_admin) OR is imported as the master_admin package by the
    fleet bundle. Avoids polluting global sys.path (which shadows other apps)."""
    import importlib
    try:
        return importlib.import_module(name)
    except ImportError:
        return importlib.import_module('master_admin.' + name)

app = Flask(__name__, template_folder='templates', static_folder='static')
# Keep the login alive for a day so we don't have to re-enter the password
# every time the browser closes.
app.permanent_session_lifetime = timedelta(days=365)

# Secret key must survive auto-reloads, or every code change kicks you out.
# Persist a generated one to a local file the first time we run.
_SECRET_FILE = ROOT / '.secret_key'
if os.environ.get('MASTER_SECRET_KEY'):
    app.secret_key = os.environ['MASTER_SECRET_KEY']
elif _SECRET_FILE.exists():
    app.secret_key = _SECRET_FILE.read_text(encoding='utf-8').strip()
else:
    app.secret_key = secrets.token_hex(32)
    _SECRET_FILE.write_text(app.secret_key, encoding='utf-8')

MASTER_PASSWORD = os.environ.get('MASTER_PASSWORD', 'lude2026')
RENDER_API_KEY = os.environ.get('RENDER_API_KEY', '')

# Google Analytics on master admin itself — set MASTER_GA_ID to enable
MASTER_GA_ID = os.environ.get('MASTER_GA_ID', 'G-H5ZV66KWSC').strip()


@app.context_processor
def _inject_master_ga():
    if not MASTER_GA_ID:
        return {'master_ga_snippet': ''}
    snippet = (
        '<!-- Google Analytics (master admin) -->\n'
        '<script async src="https://www.googletagmanager.com/gtag/js?id=' + MASTER_GA_ID + '"></script>\n'
        '<script>\n'
        'window.dataLayer = window.dataLayer || [];\n'
        'function gtag(){dataLayer.push(arguments);}\n'
        "gtag('js', new Date());\n"
        "gtag('config', '" + MASTER_GA_ID + "');\n"
        '</script>'
    )
    return {'master_ga_snippet': snippet, 'master_ga_id': MASTER_GA_ID}


# ============================================================
# Sites registry
# ============================================================

DEFAULT_SITES = [
    {
        'id': 'kai',
        'name': 'Kai Finds',
        'slug': 'kai',
        'url': 'http://127.0.0.1:5003',
        'admin_url': 'http://127.0.0.1:5003/admin',
        'admin_password': 'changeme123',
        'admin_token': '',
        'render_service_id': '',
        'agent': 'KakoBuy',
        'agent_signup': 'https://www.kakobuy.com/register?affcode=kai',
        'affiliate_code': 'kai',
        'color': '#06b6d4',
        'tag': 'template',
        'description': 'Local template — used to create new client sites',
        'product_count': 103,
    },
    {
        'id': 'jake',
        'name': 'Kakobuy.Locker',
        'slug': 'jake',
        'url': 'https://kakobuy.locker',
        'admin_url': 'https://kakobuy.locker/admin',
        'admin_password': 'jakeisgay',
        'admin_token': '',
        'render_service_id': '',
        'agent': 'KakoBuy',
        'agent_signup': '',
        'affiliate_code': 'bzaey',
        'color': '#e63946',
        'tag': 'live',
        'description': 'Live — Meta ads running',
        'product_count': 2200,
    },
    {
        'id': 'tobey',
        'name': 'JoyaFinds',
        'slug': 'tobey',
        'url': 'https://joyafinds.onrender.com',
        'admin_url': 'https://joyafinds.onrender.com/admin',
        'admin_password': 'changeme123',
        'admin_token': '',
        'render_service_id': '',
        'agent': 'JoyaGoo',
        'agent_signup': '',
        'affiliate_code': '',
        'color': '#c8e619',
        'tag': 'live',
        'description': 'JoyaGoo finds, dark theme',
        'product_count': 0,
    },
    {
        'id': 'john',
        'name': 'MrPutYouOnFinds',
        'slug': 'john',
        'url': 'https://puttingyouon.onrender.com',
        'admin_url': 'https://puttingyouon.onrender.com/admin',
        'admin_password': 'changeme123',
        'admin_token': '',
        'render_service_id': '',
        'agent': 'KakoBuy',
        'agent_signup': '',
        'affiliate_code': '',
        'color': '#a855f7',
        'tag': 'wip',
        'description': 'Work in progress',
        'product_count': 0,
    },
]


def load_sites():
    if not SITES_FILE.exists():
        save_sites(DEFAULT_SITES)
        return list(DEFAULT_SITES)
    try:
        with open(SITES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return list(DEFAULT_SITES)


def save_sites(sites):
    with open(SITES_FILE, 'w', encoding='utf-8') as f:
        json.dump(sites, f, indent=2, ensure_ascii=False)


def find_site(sid):
    for s in load_sites():
        if s.get('id') == sid:
            return s
    return None


def update_site(sid, updates):
    sites = load_sites()
    for i, s in enumerate(sites):
        if s.get('id') == sid:
            s.update(updates)
            sites[i] = s
            save_sites(sites)
            return s
    return None


# ============================================================
# Auth
# ============================================================

def is_master_admin():
    return session.get('master_admin', False)


@app.before_request
def gate():
    if request.path.startswith('/static/'):
        return
    if request.path in ('/login', '/healthz'):
        return
    if not is_master_admin():
        return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form.get('password') == MASTER_PASSWORD:
            session.permanent = True  # honour the 1-day permanent_session_lifetime
            session['master_admin'] = True
            return redirect(url_for('dashboard'))
        error = 'Wrong password'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.pop('master_admin', None)
    return redirect(url_for('login'))


@app.route('/healthz')
def healthz():
    return 'ok'


# ============================================================
# Site status + stats
# ============================================================

def check_site_status(url, timeout=2):
    """Quick HEAD ping — returns ('live', latency_ms) or ('down', 0)."""
    import requests
    if not url:
        return 'unknown', 0
    try:
        t0 = datetime.now()
        r = requests.head(url, timeout=timeout, allow_redirects=True)
        ms = int((datetime.now() - t0).total_seconds() * 1000)
        return ('live' if r.status_code < 500 else 'down'), ms
    except Exception:
        return 'down', 0


def fetch_site_stats(site, days=30):
    """Pull per-site analytics via admin API. Returns dict or None."""
    return _call_site(site, f'/admin/api/stats?days={days}', method='GET', timeout=10)


def fetch_site_products(site):
    """Pull all products from a site. Returns list or [] on failure."""
    data = _call_site(site, '/admin/api/products', method='GET', timeout=15)
    if not data:
        return []
    return data.get('products', [])


def fetch_all_sites(fn, sites, max_workers=8):
    """Run fn(site) in parallel across sites, return list of (site, result)."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    out = [None] * len(sites)
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(fn, s): i for i, s in enumerate(sites)}
        for f in as_completed(futures):
            i = futures[f]
            try:
                out[i] = (sites[i], f.result())
            except Exception:
                out[i] = (sites[i], None)
    return out


def _call_site(site, path, method='GET', json_body=None, timeout=6):
    import requests
    token = site.get('admin_token')
    base = (site.get('url') or '').rstrip('/')
    if not (token and base):
        return None
    try:
        r = requests.request(
            method,
            base + path,
            headers={'X-Admin-Token': token, 'Content-Type': 'application/json'},
            json=json_body,
            timeout=timeout,
        )
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return None


def _call_site_detailed(site, path, method='GET', json_body=None, timeout=10):
    """Like _call_site but returns (data, error). error is None on success,
    otherwise a human-readable reason so the UI can show WHY a save failed."""
    import requests
    token = site.get('admin_token')
    base = (site.get('url') or '').rstrip('/')
    if not base:
        return None, 'No site URL set — add it under Sites.'
    if not token:
        return None, 'No API key set for this site — add its admin token under Sites.'
    try:
        r = requests.request(
            method, base + path,
            headers={'X-Admin-Token': token, 'Content-Type': 'application/json'},
            json=json_body, timeout=timeout,
        )
    except requests.exceptions.Timeout:
        return None, 'Timed out reaching the site (it may be asleep — open it once, then retry).'
    except requests.exceptions.ConnectionError:
        return None, 'Could not connect — the site URL may be wrong or the site is down.'
    except Exception as e:
        return None, 'Request error: ' + str(e)[:140]
    if 200 <= r.status_code < 300:
        try:
            return r.json(), None
        except Exception:
            return {'ok': True}, None
    if r.status_code in (401, 403):
        return None, 'API key rejected (401). The token here must match the site\'s ADMIN_API_TOKEN env var.'
    if r.status_code == 404:
        return None, 'Endpoint not found (404). This site may not have this API deployed yet.'
    return None, 'Site returned HTTP ' + str(r.status_code) + '.'


# ============================================================
# Routes
# ============================================================

def _live_product_counts(sites):
    """Pull live product counts from every connected site in parallel.
    Overlays sites' `product_count` with the live number when available.
    """
    connected = [s for s in sites if s.get('admin_token')]
    if not connected:
        return sites
    results = fetch_all_sites(lambda s: fetch_site_stats(s, days=1), connected, max_workers=8)
    live_counts = {}
    for site, stats in results:
        if stats and isinstance(stats.get('total_products'), int):
            live_counts[site['id']] = stats['total_products']
    out = []
    for s in sites:
        if s['id'] in live_counts:
            s = {**s, 'product_count': live_counts[s['id']]}
        out.append(s)
    return out


@app.route('/')
def dashboard():
    sites = _live_product_counts(load_sites())
    total_products = sum(s.get('product_count', 0) for s in sites)
    live_count = sum(1 for s in sites if s.get('tag') == 'live')
    return render_template(
        'dashboard.html',
        sites=sites,
        total_products=total_products,
        live_count=live_count,
        active='dashboard',
    )


@app.route('/sites')
def sites_list():
    return render_template('sites.html', sites=_live_product_counts(load_sites()), active='sites')


@app.route('/sites/<sid>', methods=['GET', 'POST'])
def site_edit(sid):
    site = find_site(sid)
    if not site:
        return redirect(url_for('sites_list'))
    if request.method == 'POST':
        updates = {}
        for k in ('name', 'url', 'admin_url', 'admin_password', 'admin_token',
                  'render_service_id', 'agent', 'agent_signup',
                  'affiliate_code', 'color', 'tag', 'description',
                  'ga_measurement_id', 'ga_property_id'):
            if k in request.form:
                updates[k] = request.form.get(k, '').strip()
        try:
            updates['product_count'] = int(request.form.get('product_count') or 0)
        except ValueError:
            pass
        update_site(sid, updates)
        return redirect(url_for('site_edit', sid=sid, saved=1))
    return render_template('site_edit.html', site=site, active='sites',
                            saved=request.args.get('saved'))


@app.route('/sites/new', methods=['GET', 'POST'])
def site_new():
    if request.method == 'POST':
        # MVP: just append to sites.json. Actual provisioning via create_site.py
        # is a separate "Deploy" button (Phase 2).
        sid = (request.form.get('slug') or '').strip().lower()
        if not sid:
            return render_template('site_new.html', error='Slug is required', active='sites')
        sites = load_sites()
        if any(s.get('id') == sid for s in sites):
            return render_template('site_new.html', error='Slug already exists', active='sites')
        new = {
            'id': sid,
            'slug': sid,
            'name': request.form.get('name', sid.title()).strip(),
            'url': request.form.get('url', '').strip(),
            'admin_url': '',
            'admin_password': '',
            'admin_token': '',
            'render_service_id': '',
            'agent': request.form.get('agent', 'KakoBuy').strip(),
            'agent_signup': '',
            'affiliate_code': request.form.get('affiliate_code', '').strip(),
            'color': request.form.get('color', '#06b6d4').strip(),
            'tag': 'wip',
            'description': '',
            'product_count': 0,
        }
        if new['url']:
            new['admin_url'] = new['url'].rstrip('/') + '/admin'
        sites.append(new)
        save_sites(sites)
        return redirect(url_for('site_edit', sid=sid, saved=1))
    return render_template('site_new.html', active='sites')


@app.route('/sites/<sid>/delete', methods=['POST'])
def site_delete(sid):
    sites = [s for s in load_sites() if s.get('id') != sid]
    save_sites(sites)
    return jsonify({'ok': True})


@app.route('/sites/<sid>/status')
def site_status(sid):
    """AJAX endpoint — public-URL ping. Returns live/down + latency."""
    site = find_site(sid)
    if not site:
        return jsonify({'error': 'not found'}), 404
    status, latency = check_site_status(site.get('url', ''))
    return jsonify({'id': sid, 'status': status, 'latency_ms': latency})


@app.route('/sites/<sid>/test-api')
def site_test_api(sid):
    """Ping the admin API with the stored token — returns rich info for the site_edit UI."""
    site = find_site(sid)
    if not site:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    if not site.get('admin_token'):
        return jsonify({'ok': False, 'error': 'no token saved'}), 200
    if not site.get('url'):
        return jsonify({'ok': False, 'error': 'no URL'}), 200
    from datetime import datetime
    t0 = datetime.now()
    resp = _call_site(site, '/admin/api/ping', timeout=6)
    latency = int((datetime.now() - t0).total_seconds() * 1000)
    if not resp:
        return jsonify({'ok': False, 'error': 'API unreachable or returned non-200', 'latency_ms': latency})
    return jsonify({
        'ok': True,
        'latency_ms': latency,
        'site_name': resp.get('site'),
        'agent': resp.get('agent'),
        'token_valid': resp.get('token_valid'),
    })


@app.route('/products')
def products():
    """Unified product manager — pulls from every site with an API token, in parallel."""
    sites = load_sites()
    connected = [s for s in sites if s.get('admin_token')]
    results = fetch_all_sites(fetch_site_products, connected)
    all_products = []
    reachable_sites = []
    unreachable_sites = []
    for site, prods in results:
        if prods is None or len(prods) == 0:
            # Could be "no products" (still reachable) or unreachable — distinguish via ping
            ping = _call_site(site, '/admin/api/ping', timeout=4)
            if ping and ping.get('token_valid'):
                reachable_sites.append(site)
                continue
            unreachable_sites.append(site)
            continue
        reachable_sites.append(site)
        for p in prods:
            p['_site_id'] = site['id']
            p['_site_name'] = site['name']
            p['_site_color'] = site['color']
            all_products.append(p)

    # Filters
    q = (request.args.get('q') or '').lower()
    site_filter = request.args.get('site', '')
    cat_filter = request.args.get('cat', '')
    if q:
        all_products = [p for p in all_products if q in (p.get('name','').lower() + ' ' + p.get('tags','').lower() + ' ' + p.get('seller','').lower())]
    if site_filter:
        all_products = [p for p in all_products if p['_site_id'] == site_filter]
    if cat_filter:
        all_products = [p for p in all_products if p.get('category') == cat_filter]

    # All categories across all sites
    cats = sorted({p.get('category','') for p in all_products if p.get('category')})

    # Pagination
    page = max(1, request.args.get('page', 1, type=int))
    per_page = 50
    total = len(all_products)
    total_pages = max(1, (total + per_page - 1) // per_page)
    start = (page - 1) * per_page
    paged = all_products[start:start + per_page]

    sites_with_tokens = connected
    sites_without_tokens = [s for s in sites if not s.get('admin_token')]

    return render_template('products.html',
        sites=sites,
        sites_with_tokens=sites_with_tokens,
        sites_without_tokens=sites_without_tokens,
        unreachable_sites=unreachable_sites,
        products=paged,
        total=total,
        page=page,
        total_pages=total_pages,
        cats=cats,
        q=q,
        site_filter=site_filter,
        cat_filter=cat_filter,
        active='products')


@app.route('/products/push', methods=['POST'])
def products_push():
    """Bulk-push a product to multiple target sites."""
    data = request.get_json(silent=True) or {}
    payload = data.get('product') or {}
    targets = data.get('targets') or []
    if not payload or not targets:
        return jsonify({'error': 'product and targets required'}), 400
    sites = {s['id']: s for s in load_sites()}
    results = {}
    for tid in targets:
        site = sites.get(tid)
        if not site:
            results[tid] = {'ok': False, 'error': 'site not found'}
            continue
        # Strip site-specific metadata before pushing
        body = {k: v for k, v in payload.items() if not k.startswith('_')}
        body.pop('id', None)  # let target generate a fresh ID
        resp = _call_site(site, '/admin/api/products', method='POST', json_body=body)
        results[tid] = {'ok': bool(resp and resp.get('ok')), 'data': resp}
    return jsonify({'ok': True, 'results': results})


@app.route('/products/scrape', methods=['POST'])
def products_scrape():
    """Scrape a Weidian/Taobao/1688 (or agent-wrapped) link for the add-product composer.
    Returns scrapable fields (name, base + per-style prices, variants, gallery) plus a
    manifest of fields the operator must add by hand (QC photos, sizes, batch, etc.)."""
    scrape_product = _sibling('scrape_product')
    data = request.get_json(silent=True) or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'ok': False, 'error': 'url required'}), 400
    try:
        return jsonify(scrape_product.scrape(url))
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)[:200]}), 200


@app.route('/products/<site_id>/new', methods=['POST'])
def products_create_one(site_id):
    """Create a brand-new product on the origin site (works on any domain)."""
    sites = {s['id']: s for s in load_sites()}
    site = sites.get(site_id)
    if not site:
        return jsonify({'error': 'site not found'}), 404
    body = request.get_json(silent=True) or {}
    body = {k: v for k, v in body.items() if not k.startswith('_')}
    if not (body.get('name') or '').strip():
        return jsonify({'ok': False, 'error': 'name required'}), 400
    data, err = _call_site_detailed(site, '/admin/api/products', method='POST', json_body=body)
    if err:
        return jsonify({'ok': False, 'error': err}), 502
    return jsonify(data or {'ok': True})


@app.route('/search/<site_id>')
def search_proxy(site_id):
    """Proxy a catalogue search to the origin site (for the Best-Selling picker)."""
    from urllib.parse import quote
    site = next((s for s in load_sites() if s['id'] == site_id), None)
    if not site:
        return jsonify({'results': []})
    q = (request.args.get('q') or '').strip()
    resp = _call_site(site, '/api/search?limit=12&q=' + quote(q))
    return jsonify(resp or {'results': []})


@app.route('/upload/<site_id>/<pid>', methods=['POST'])
def upload_proxy(site_id, pid):
    """Forward a multipart image upload to the origin site (token-authed)."""
    import requests
    site = next((s for s in load_sites() if s['id'] == site_id), None)
    if not site or not site.get('url'):
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'no file'}), 400
    f = request.files['file']
    base = site['url'].rstrip('/')
    primary = request.args.get('primary', '1')
    try:
        r = requests.post(base + '/admin/products/upload-image/' + pid + '?primary=' + primary,
                          headers={'X-Admin-Token': site.get('admin_token', '')},
                          files={'file': (f.filename, f.stream, f.mimetype or 'application/octet-stream')},
                          timeout=40)
        try:
            return jsonify(r.json()), r.status_code
        except Exception:
            return jsonify({'ok': False, 'error': 'upload failed (HTTP ' + str(r.status_code) + ')'}), 502
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)[:160]}), 502


@app.route('/products/<site_id>/gallery/<pid>', methods=['POST'])
def products_gallery(site_id, pid):
    """Proxy a product's manual photo-gallery update to the origin site."""
    sites = {s['id']: s for s in load_sites()}
    site = sites.get(site_id)
    if not site:
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    body = request.get_json(silent=True) or {}
    data, err = _call_site_detailed(site, '/admin/products/gallery/' + pid, method='POST', json_body=body)
    if err:
        return jsonify({'ok': False, 'error': err}), 502
    return jsonify(data or {'ok': True})


@app.route('/products/<site_id>/bulk', methods=['POST'])
def products_bulk(site_id):
    """Proxy a batch of product edits to the origin site in one request."""
    sites = {s['id']: s for s in load_sites()}
    site = sites.get(site_id)
    if not site:
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    body = request.get_json(silent=True) or {}
    data, err = _call_site_detailed(site, '/admin/api/products/bulk', method='POST', json_body=body, timeout=20)
    if err:
        return jsonify({'ok': False, 'error': err}), 502
    return jsonify(data or {'ok': True})


@app.route('/products/<site_id>/<pid>', methods=['GET'])
def products_get_one(site_id, pid):
    """Fetch the raw product row from the origin site (for duplicate / undo)."""
    site = next((s for s in load_sites() if s['id'] == site_id), None)
    if not site:
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    return jsonify(_call_site(site, '/admin/api/products/' + pid) or {'ok': False, 'error': 'unreachable'})


@app.route('/products/<site_id>/<pid>', methods=['PUT', 'POST'])
def products_update_one(site_id, pid):
    """Proxy an inline product edit to the origin site (works on any domain)."""
    sites = {s['id']: s for s in load_sites()}
    site = sites.get(site_id)
    if not site:
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    body = request.get_json(silent=True) or {}
    body = {k: v for k, v in body.items() if not k.startswith('_')}
    data, err = _call_site_detailed(site, f'/admin/api/products/{pid}', method='PUT', json_body=body)
    if err:
        return jsonify({'ok': False, 'error': err}), 502
    return jsonify(data or {'ok': True})


@app.route('/products/<site_id>/<pid>', methods=['DELETE'])
def products_delete_one(site_id, pid):
    sites = {s['id']: s for s in load_sites()}
    site = sites.get(site_id)
    if not site:
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    data, err = _call_site_detailed(site, f'/admin/api/products/{pid}', method='DELETE')
    if err:
        return jsonify({'ok': False, 'error': err}), 502
    return jsonify(data or {'ok': True})


# === Site Content editor (hero, nav, footer, branding, theme) ================
@app.route('/content')
def content_index():
    # The side-panel content form was superseded by the immersive Studio.
    return redirect('/studio')


@app.route('/content/<site_id>')
def content_edit(site_id):
    return redirect('/studio/' + site_id)


@app.route('/content/<site_id>', methods=['POST', 'PUT'])
def content_save(site_id):
    sites = {s['id']: s for s in load_sites()}
    site = sites.get(site_id)
    if not site:
        return jsonify({'error': 'site not found'}), 404
    body = request.get_json(silent=True) or {}
    body = {k: v for k, v in body.items() if not k.startswith('_')}
    data, err = _call_site_detailed(site, '/admin/api/settings', method='PUT', json_body=body)
    if err:
        return jsonify({'ok': False, 'error': err}), 502
    return jsonify(data or {'ok': True})


# === Immersive visual Studio (full-screen in-place editor) ===================
@app.route('/studio')
def studio_index():
    sites = load_sites()
    connected = [s for s in sites if s.get('url') and s.get('admin_token')]
    # Prefer the in-bundle same-origin site (ategoat) which the in-place editor fully
    # supports; fall back to the first connected site, then anything.
    pref = next((s for s in connected if s['id'] == 'ategoat'), None) or (connected[0] if connected else None) or (sites[0] if sites else None)
    if pref:
        return redirect('/studio/' + pref['id'])
    return render_template('studio.html', site=None, sites=sites, active='content')


@app.route('/studio/<site_id>')
def studio(site_id):
    sites = load_sites()
    site = next((s for s in sites if s['id'] == site_id), None)
    if not site:
        return redirect('/studio')
    return render_template('studio.html', site=site, sites=sites, active='content')


# --- category proxies (used by the in-place editor) ---
@app.route('/categories/<site_id>')
def categories_list(site_id):
    site = next((s for s in load_sites() if s['id'] == site_id), None)
    if not site:
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    return jsonify(_call_site(site, '/admin/api/categories') or {'ok': False, 'categories': []})


@app.route('/categories/<site_id>', methods=['POST'])
def categories_add(site_id):
    site = next((s for s in load_sites() if s['id'] == site_id), None)
    if not site:
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    data, err = _call_site_detailed(site, '/admin/api/categories', method='POST', json_body=request.get_json(silent=True) or {})
    return (jsonify({'ok': False, 'error': err}), 502) if err else jsonify(data or {'ok': True})


@app.route('/categories/<site_id>/reorder', methods=['POST'])
def categories_reorder(site_id):
    site = next((s for s in load_sites() if s['id'] == site_id), None)
    if not site:
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    data, err = _call_site_detailed(site, '/admin/api/categories/reorder', method='POST', json_body=request.get_json(silent=True) or {})
    return (jsonify({'ok': False, 'error': err}), 502) if err else jsonify(data or {'ok': True})


@app.route('/categories/<site_id>/<slug>', methods=['PATCH', 'PUT'])
def categories_update(site_id, slug):
    site = next((s for s in load_sites() if s['id'] == site_id), None)
    if not site:
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    data, err = _call_site_detailed(site, '/admin/api/categories/' + slug, method='PATCH', json_body=request.get_json(silent=True) or {})
    return (jsonify({'ok': False, 'error': err}), 502) if err else jsonify(data or {'ok': True})


@app.route('/categories/<site_id>/<slug>', methods=['DELETE'])
def categories_delete(site_id, slug):
    site = next((s for s in load_sites() if s['id'] == site_id), None)
    if not site:
        return jsonify({'ok': False, 'error': 'site not found'}), 404
    data, err = _call_site_detailed(site, '/admin/api/categories/' + slug, method='DELETE', json_body=request.get_json(silent=True) or {})
    return (jsonify({'ok': False, 'error': err}), 502) if err else jsonify(data or {'ok': True})


@app.route('/analytics')
def analytics():
    """Aggregated analytics across all sites that expose an API token."""
    sites = load_sites()
    days = request.args.get('days', 30, type=int)
    connected = [s for s in sites if s.get('admin_token')]

    # Current period
    results = fetch_all_sites(lambda s: fetch_site_stats(s, days=days), connected)
    stats_by_id = {site['id']: stats for site, stats in results}
    # Previous period (e.g. days 30-60 ago for trend comparison)
    prev_results = fetch_all_sites(lambda s: fetch_site_stats(s, days=days * 2), connected)
    prev_stats_by_id = {site['id']: stats for site, stats in prev_results}

    per_site = []
    aggregate = {'total_clicks': 0, 'unique_visitors': 0, 'signup_clicks': 0,
                 'total_products': 0, 'top_products': {}, 'top_categories': {}, 'daily': {}}
    daily_by_site = {}  # site_id -> {day: clicks}
    prev_aggregate = {'total_clicks': 0, 'unique_visitors': 0, 'signup_clicks': 0}

    for s in sites:
        if not s.get('admin_token'):
            per_site.append({'site': s, 'stats': None, 'delta': None})
            continue
        stats = stats_by_id.get(s['id'])
        prev2x = prev_stats_by_id.get(s['id']) or {}

        # Compute previous-period (the older half of the 2x window)
        prev_clicks = max(0, prev2x.get('total_clicks', 0) - (stats.get('total_clicks', 0) if stats else 0))
        prev_signups = max(0, prev2x.get('signup_clicks', 0) - (stats.get('signup_clicks', 0) if stats else 0))
        prev_visitors = max(0, prev2x.get('unique_visitors', 0) - (stats.get('unique_visitors', 0) if stats else 0))

        delta = None
        if stats:
            def pct(now, prev):
                if not prev:
                    return None if not now else 100.0
                return ((now - prev) / prev) * 100
            delta = {
                'clicks_pct':   pct(stats.get('total_clicks', 0), prev_clicks),
                'visitors_pct': pct(stats.get('unique_visitors', 0), prev_visitors),
                'signups_pct':  pct(stats.get('signup_clicks', 0), prev_signups),
            }

        per_site.append({'site': s, 'stats': stats, 'delta': delta})
        if not stats:
            continue
        aggregate['total_clicks']    += stats.get('total_clicks', 0)
        aggregate['unique_visitors'] += stats.get('unique_visitors', 0)
        aggregate['signup_clicks']   += stats.get('signup_clicks', 0)
        aggregate['total_products']  += stats.get('total_products', 0)
        prev_aggregate['total_clicks']    += prev_clicks
        prev_aggregate['unique_visitors'] += prev_visitors
        prev_aggregate['signup_clicks']   += prev_signups

        site_daily = {}
        for tp in (stats.get('top_products') or []):
            aggregate['top_products'][tp['product_name']] = aggregate['top_products'].get(tp['product_name'], 0) + tp['clicks']
        for tc in (stats.get('top_categories') or []):
            aggregate['top_categories'][tc['category']] = aggregate['top_categories'].get(tc['category'], 0) + tc['clicks']
        for d in (stats.get('daily') or []):
            day = d.get('day')
            if day:
                aggregate['daily'][day] = aggregate['daily'].get(day, 0) + d.get('clicks', 0)
                site_daily[day] = d.get('clicks', 0)
        daily_by_site[s['id']] = site_daily

    top_products  = sorted(aggregate['top_products'].items(),  key=lambda x: -x[1])[:15]
    top_categories = sorted(aggregate['top_categories'].items(), key=lambda x: -x[1])[:15]
    daily = sorted(aggregate['daily'].items())
    conversion = (aggregate['signup_clicks'] / aggregate['total_clicks'] * 100) if aggregate['total_clicks'] else 0
    prev_conversion = (prev_aggregate['signup_clicks'] / prev_aggregate['total_clicks'] * 100) if prev_aggregate['total_clicks'] else 0

    def pct_change(now, prev):
        if not prev:
            return None if not now else 100.0
        return round(((now - prev) / prev) * 100, 1)

    deltas = {
        'clicks':     pct_change(aggregate['total_clicks'], prev_aggregate['total_clicks']),
        'visitors':   pct_change(aggregate['unique_visitors'], prev_aggregate['unique_visitors']),
        'signups':    pct_change(aggregate['signup_clicks'], prev_aggregate['signup_clicks']),
        'conversion': round(conversion - prev_conversion, 1),
    }

    # Insights (auto-flagged issues for the executive view)
    insights = []
    for row in per_site:
        s = row['site']; st = row['stats']
        if not st:
            if s.get('admin_token'):
                insights.append({'level': 'error', 'site': s['name'], 'msg': 'API unreachable — site may be down or token invalid'})
            continue
        if st.get('total_products', 0) > 0 and st.get('total_clicks', 0) == 0:
            insights.append({'level': 'warn', 'site': s['name'], 'msg': f"{st['total_products']:,} products but zero clicks — needs traffic"})
        # NOTE: signup completion happens on the agent's site — we can't track it,
        # so we don't generate any "no signups" insight. Keeping rules to outbound
        # click activity only.

    return render_template('analytics.html',
        sites=sites,
        days=days,
        per_site=per_site,
        agg=aggregate,
        prev_agg=prev_aggregate,
        deltas=deltas,
        top_products=top_products,
        top_categories=top_categories,
        daily=daily,
        daily_by_site=daily_by_site,
        conversion=round(conversion, 1),
        prev_conversion=round(prev_conversion, 1),
        insights=insights,
        active='analytics')


@app.route('/export/<kind>')
def export(kind):
    """Export master-admin data as CSV or Excel.

    Kinds: products | sites | analytics-daily | analytics-products | analytics-categories
    Format: ?format=xlsx (default) or ?format=csv
    """
    from io import BytesIO, StringIO
    import csv as csv_mod
    from flask import make_response

    fmt = (request.args.get('format') or 'xlsx').lower()
    sites = load_sites()

    sheets = {}  # sheet name → (headers, rows)

    if kind == 'sites':
        sheets['Sites'] = (
            ['ID', 'Name', 'URL', 'Agent', 'Affcode', 'Color', 'Status', 'Products', 'Description'],
            [[s['id'], s['name'], s.get('url',''), s.get('agent',''), s.get('affiliate_code',''),
              s.get('color',''), s.get('tag',''), s.get('product_count', 0), s.get('description','')]
             for s in sites],
        )
    elif kind == 'products':
        connected = [s for s in sites if s.get('admin_token')]
        results = fetch_all_sites(fetch_site_products, connected)
        rows = []
        MAX_ROWS = 60000  # hard cap so a multi-site export can't OOM the instance
        truncated = False
        for site, prods in results:
            if not prods:
                continue
            for p in prods:
                if len(rows) >= MAX_ROWS:
                    truncated = True
                    break
                rows.append([
                    site['name'], p.get('id', ''), p.get('name', ''), p.get('category', ''),
                    p.get('price', ''), p.get('seller', ''), p.get('url', ''), p.get('image', ''),
                    'YES' if p.get('featured') else '', p.get('tags', '')
                ])
            if truncated:
                break
        if truncated:
            rows.append(['(truncated at %d rows — use CSV for the full export)' % MAX_ROWS, '', '', '', '', '', '', '', '', ''])
        sheets['Products'] = (
            ['Site', 'ID', 'Name', 'Category', 'Price USD', 'Seller', 'Product URL', 'Image', 'Featured', 'Tags'],
            rows,
        )
    elif kind in ('analytics-daily', 'analytics-products', 'analytics-categories', 'analytics-all'):
        days = request.args.get('days', 30, type=int)
        connected = [s for s in sites if s.get('admin_token')]
        results = fetch_all_sites(lambda s: fetch_site_stats(s, days=days), connected)

        if kind in ('analytics-daily', 'analytics-all'):
            daily_rows = []
            for site, stats in results:
                if not stats: continue
                for d in (stats.get('daily') or []):
                    daily_rows.append([site['name'], d.get('day'), d.get('clicks', 0), d.get('visitors', 0)])
            sheets['Daily Activity'] = (['Site', 'Date', 'Clicks', 'Visitors'], daily_rows)

        if kind in ('analytics-products', 'analytics-all'):
            prod_rows = []
            for site, stats in results:
                if not stats: continue
                for rank, p in enumerate(stats.get('top_products') or [], 1):
                    prod_rows.append([site['name'], rank, p.get('product_name'), p.get('clicks', 0)])
            sheets['Top Products'] = (['Site', 'Rank', 'Product', 'Clicks'], prod_rows)

        if kind in ('analytics-categories', 'analytics-all'):
            cat_rows = []
            for site, stats in results:
                if not stats: continue
                for rank, c in enumerate(stats.get('top_categories') or [], 1):
                    cat_rows.append([site['name'], rank, c.get('category'), c.get('clicks', 0)])
            sheets['Top Categories'] = (['Site', 'Rank', 'Category', 'Clicks'], cat_rows)

        if kind == 'analytics-all':
            # Fleet summary sheet
            summary_rows = []
            for site, stats in results:
                if not stats:
                    summary_rows.append([site['name'], '—', '—', '—', '—'])
                    continue
                summary_rows.append([site['name'],
                                     stats.get('total_clicks', 0),
                                     stats.get('unique_visitors', 0),
                                     stats.get('total_products', 0),
                                     stats.get('categories', 0)])
            sheets = {'Fleet Summary': (['Site', 'Clicks', 'Visitors', 'Products', 'Categories'], summary_rows), **sheets}
    else:
        return jsonify({'error': 'unknown kind'}), 400

    base_name = f'master-{kind}-{datetime.now().strftime("%Y%m%d-%H%M")}'

    if fmt == 'csv':
        # Stream row-by-row so even a huge fleet export uses ~no memory.
        from flask import Response, stream_with_context
        sheet_name, (headers, rows) = next(iter(sheets.items()))

        def _gen():
            buf = StringIO(); w = csv_mod.writer(buf)
            w.writerow(headers); yield buf.getvalue(); buf.seek(0); buf.truncate(0)
            for r in rows:
                w.writerow(r); yield buf.getvalue(); buf.seek(0); buf.truncate(0)
        resp = Response(stream_with_context(_gen()), mimetype='text/csv; charset=utf-8')
        resp.headers['Content-Disposition'] = f'attachment; filename="{base_name}.csv"'
        return resp

    # XLSX — write_only mode keeps memory bounded (no per-cell style objects held
    # for every row, which is what OOM'd the instance on big product exports).
    try:
        import openpyxl
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({'error': 'openpyxl not installed — pip install openpyxl'}), 500
    try:
        wb = openpyxl.Workbook(write_only=True)
        hfont = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill('solid', fgColor='06B6D4')
        for sheet_name, (headers, rows) in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            hdr = []
            for h in headers:
                c = WriteOnlyCell(ws, value=h); c.font = hfont; c.fill = hfill; hdr.append(c)
            ws.append(hdr)
            for r in rows:
                ws.append(r)
        if not wb.sheetnames:
            ws = wb.create_sheet('Empty'); ws.append(['No data'])
        buf = BytesIO()
        wb.save(buf)
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="{base_name}.xlsx"'
        return resp
    except Exception as e:
        return jsonify({'error': 'Export failed (try CSV for very large catalogues): ' + str(e)[:150]}), 500


@app.route('/api/ga/<sid>')
def api_ga(sid):
    """Pull live GA data for a site. Returns null if not configured."""
    ga_client = _sibling('ga_client')
    site = find_site(sid)
    if not site:
        return jsonify({'error': 'site not found'}), 404
    pid = site.get('ga_property_id')
    if not pid:
        return jsonify({'configured': False, 'reason': 'no_property_id', 'site': sid})
    if not ga_client.is_configured():
        return jsonify({'configured': False, 'reason': 'no_credentials', 'site': sid})
    days = request.args.get('days', 30, type=int)
    data = ga_client.get_stats(pid, days=days)
    if data is None:
        return jsonify({'configured': False, 'reason': 'client_init_failed', 'site': sid})
    if isinstance(data, dict) and 'error' in data:
        return jsonify({'configured': True, 'error': data['error'], 'site': sid})
    return jsonify({'configured': True, 'site': sid, 'days': days, 'data': data})


@app.route('/ga/setup')
def ga_setup():
    """Walks the user through getting GA Data API credentials."""
    ga_client = _sibling('ga_client')
    return render_template('ga_setup.html',
        active='analytics',
        is_configured=ga_client.is_configured(),
        sites=load_sites(),
    )


TEMPLATES_REGISTRY = [
    {
        'slug': 'volume',
        'name': 'Volume Studio',
        'status': 'live',
        'description': 'Magazine/zine catalogue template — warm cream-paper aesthetic, Playfair Display serif italic accents, masthead dateline bar, TOC-numbered sections, outfit-bundle cards (3-image stack + piece count + total price), and a printed-magazine feel. Inspired by finds.org with editorial polish + heavy print-design treatment. Highly customizable via env vars (issue_number, hero_tagline, hero_sub_tagline).',
        'color': '#dc2626',
        'theme': 'Light · magazine/zine',
        'natural_theme': 'light',
        'natural_font': 'playfair',
        'fonts': 'Playfair Display italic · Inter · JetBrains Mono',
        'features': ['Top dateline bar with live date', 'Masthead nav with issue number', 'TOC-style section numbering', 'Stacked 2-line italic serif hero', 'Marquee ticker under hero', 'Featured editorial cover story', 'Outfit bundle cards (3-image collage + piece count + total)', 'Cream paper background with subtle texture', 'Square corners + 2px section borders', 'Customizable issue_number / hero_tagline / hero_sub_tagline env vars'],
        'preview_url': '/volume',
        'folder': 'volume',
        'product_count_label': '100+',
    },
    {
        'slug': 'future',
        'name': 'Future Finds',
        'status': 'live',
        'description': 'Sci-fi mission-control template — deep dark bg, electric cyan accent, monospace data labels, terminal-style readouts, scanline overlay, live UTC clock. Built for clients who want a Bloomberg-meets-Apple-Vision-Pro vibe.',
        'color': '#00f5d4',
        'theme': 'Dark · sci-fi terminal',
        'natural_theme': 'dark',
        'natural_font': 'space',
        'fonts': 'Space Grotesk · JetBrains Mono · Inter',
        'features': ['Live UTC clock in nav', 'Terminal-style spec sheet readouts', 'Scanline overlay', 'Pulsing status dot indicators', 'Monospace data labels everywhere', 'Tabular numerals', 'Cyan glow on hover', 'Telemetry strip marquee', 'Numerical product indices'],
        'preview_url': '/future',
        'folder': 'future',
        'product_count_label': '100+',
    },
    {
        'slug': 'maywood',
        'name': 'Maywood Sheets',
        'status': 'live',
        'description': 'Refined editorial finds template — light-by-default with serif italic display, doppel.fit-style cards (1:1 rounded with variant cluster pill), and finds.org-inspired layout (brand strip marquee, outfit bundles, dense category nav, Discord CTA). Built for clients who want editorial polish over neon energy.',
        'color': '#0d9488',
        'theme': 'Light · refined editorial',
        'natural_theme': 'light',
        'natural_font': 'inter',
        'fonts': 'Instrument Serif italic · Inter · JetBrains Mono',
        'features': ['Two-line serif italic hero', 'Variant cluster pill on cards', 'Light/dark/mono themes', 'Outfit bundle rails', 'Brand strip marquee', '1:1 rounded cards', 'Theme toggle', 'Discord banner', 'Italic accent typography'],
        'preview_url': '/maywood',
        'folder': 'maywood',
        'product_count_label': '100+',
    },
    {
        'slug': 'kai',
        'name': 'Kai Finds',
        'status': 'live',
        'description': 'Classic dark fashion-finds site with cyan brand color. Hero ticker, product grid, link converter, tutorial, side-widget. Battle-tested with 4 deployed clients.',
        'color': '#06b6d4',
        'theme': 'Dark · streetwear',
        'natural_theme': 'dark',
        'natural_font': 'anton',
        'fonts': 'Anton · Inter · Space Grotesk',
        'features': ['Product detail modal', 'Currency switcher', 'EN/中文', 'Auto-tags + fuzzy search', 'Featured pins', 'Drag reorder', '3-view admin', 'Side widget'],
        'preview_url': '/kai',
        'folder': 'kai',
        'product_count_label': '100+',
    },
    {
        'slug': 'editorial',
        'name': 'Editorial House',
        'status': 'live',
        'description': 'High-end fashion editorial — warm ivory paper, deep editorial red accent, Playfair Display serif headlines with DM Serif italic counterpoint. Split cover hero with masthead dateline, marquee ticker, magazine-style spreads, pull quotes, lookbook grid, table-of-contents tile index, editors note, and a high-contrast subscription card. Built for clients selling curation and taste — quiet authority, considered restraint, no neon.',
        'color': '#8a1c1c',
        'theme': 'Light · editorial',
        'fonts': 'Playfair Display · DM Serif Display italic · Cormorant Garamond · Inter · JetBrains Mono',
        'features': ['Top dateline strip with issue + season', 'Masthead nav with serif logo + italic accent', 'Split editorial hero (big serif + cover image)', 'Marquee ticker under hero', 'Featured selection (8 cards, № 01 indices)', 'Blown-up Feature spread (single piece)', 'Pull quote section', 'Latest entries conveyor', 'Table-of-contents tile index', 'Lookbook 3-col spread', 'Editors Note colophon', 'Hard-contrast CTA strip', 'Subscription popup as house note', 'Light + dark + mono themes', 'Customizable issue_label / hero_title / editor_name'],
        'preview_url': '/editorial',
        'folder': 'editorial',
        'product_count_label': '100+',
        'natural_theme': 'light',
        'natural_font': 'playfair',
    },
    {
        'slug': 'minimal',
        'name': 'Minimalist Template',
        'status': 'live',
        'description': 'Pure monochrome minimalism. Single Inter font in all weights, mono captions, square corners, fine 1px borders, generous whitespace. No accents, no gradients, no decoration — the product does the work, not the chrome around it.',
        'color': '#0a0a0a',
        'theme': 'Light · monochrome',
        'natural_theme': 'light',
        'natural_font': 'inter',
        'fonts': 'Inter · JetBrains Mono',
        'features': ['Pure monochrome', 'Inter-only typography', 'Square corners (0 radius)', 'Index-numbered sections', 'Generous whitespace', 'Mono-caption labels', 'Light/dark/mono themes'],
        'preview_url': '/minimal',
        'folder': 'minimal',
        'product_count_label': '100+',
    },
    {
        'slug': 'vaporwave',
        'name': 'Vaporwave Y2K',
        'status': 'coming-soon',
        'description': 'Pink/purple gradient mesh, holographic glassmorphism, retro Y2K vibe. Cards float with subtle iridescent shimmer.',
        'color': '#ec4899',
        'theme': 'Dark · Y2K aesthetic',
        'fonts': 'Syne · Space Mono',
        'features': ['Gradient mesh backgrounds', 'Holographic cards', 'CRT scan-line effects', 'Vapor product names', 'Glassmorphism nav'],
        'preview_url': '',
        'folder': '',
    },
    {
        'slug': 'brutalist',
        'name': 'Brutalist Grid',
        'status': 'coming-soon',
        'description': 'Stark black/white, raw HTML aesthetic, sharp grids, no rounded corners. Cortis-style. For audiences that find polish suspicious.',
        'color': '#fff',
        'theme': 'Mono · brutalist',
        'fonts': 'JetBrains Mono · system',
        'features': ['Mono everywhere', 'Sharp 0-radius corners', 'Heavy borders', '4xN grid layouts', 'Marquee tickers'],
        'preview_url': '',
        'folder': '',
    },
    {
        'slug': 'tile-market',
        'name': 'Tile Market',
        'status': 'coming-soon',
        'description': 'Pinterest/Etsy-style dense tile grid. Pictures are king. Hover for info, click for modal. Best for sites with strong product imagery.',
        'color': '#22c55e',
        'theme': 'Dark · visual-first',
        'fonts': 'DM Sans · Inter',
        'features': ['Masonry product grid', 'Image-zoom on hover', 'Quick-view modal', 'Infinite scroll', 'Filter chips'],
        'preview_url': '',
        'folder': '',
    },
    {
        'slug': 'storytime',
        'name': 'Storytime Drops',
        'status': 'coming-soon',
        'description': 'Vertical scroll narrative — each product gets a full-bleed hero with story. Slow, intentional, premium feel. Inspired by Apple product pages.',
        'color': '#8b5cf6',
        'theme': 'Dark · narrative',
        'fonts': 'Anton · Inter',
        'features': ['Full-bleed hero per product', 'Scroll-snap sections', 'Parallax images', 'Story copy per drop', 'CTA reveals'],
        'preview_url': '',
        'folder': '',
    },
    {
        'slug': 'terminal',
        'name': 'Terminal Mode',
        'status': 'live',
        'description': 'Hacker-aesthetic phosphor CRT terminal. JetBrains Mono everywhere, $-prompt UI elements, scanline overlay, boot-sequence hero. For audiences that like a sci-fi CLI vibe.',
        'color': '#22c55e',
        'theme': 'Dark · CLI',
        'natural_theme': 'dark',
        'natural_font': 'jetbrains',
        'fonts': 'JetBrains Mono',
        'features': ['Mono UI', 'CRT scanlines', '$-prompt nav', 'Boot-sequence hero', 'Indexed product packets', 'Phosphor-green palette'],
        'preview_url': '/terminal',
        'folder': 'terminal',
    },
    {
        'slug': 'ategoat',
        'name': 'AteGoat',
        'status': 'live',
        'description': 'Maywood-template site populated with the full 9,400+-product ategoat.com catalogue (scraped via their wiligoods WordPress API). Light/refined editorial cards, teal accent, KakoBuy affiliate ready. Use as a starting point for a client who wants a large finds catalogue out of the box.',
        'color': '#0d9488',
        'theme': 'Light · refined editorial',
        'natural_theme': 'light',
        'natural_font': 'inter',
        'fonts': 'Instrument Serif italic · Inter · JetBrains Mono',
        'features': ['9,400+ pre-scraped products', '13 categories', 'KakoBuy ready', 'Maywood card style', 'Italic serif hero', 'Customizable via env vars'],
        'preview_url': '/ategoat',
        'folder': 'ategoat',
        'product_count_label': '9,400+',
    },
]


@app.route('/templates')
def templates_list():
    """Template library — browse available site designs."""
    return render_template(
        'templates_list.html',
        templates=TEMPLATES_REGISTRY,
        active='templates',
    )


@app.route('/templates/<slug>')
def template_detail(slug):
    """Single template page with full details + use-this-template CTA."""
    tpl = next((t for t in TEMPLATES_REGISTRY if t['slug'] == slug), None)
    if not tpl:
        return redirect(url_for('templates_list'))
    return render_template('template_detail.html', tpl=tpl, active='templates')


@app.route('/templates/<slug>/studio')
def template_studio(slug):
    """Interactive preview studio — clients sandbox brand name / color / theme / logo on a real template."""
    tpl = next((t for t in TEMPLATES_REGISTRY if t['slug'] == slug), None)
    if not tpl:
        return redirect(url_for('templates_list'))
    if tpl['status'] != 'live' or not tpl.get('preview_url'):
        return redirect(url_for('template_detail', slug=slug))
    return render_template('template_studio.html', tpl=tpl, active='templates')


@app.route('/tools')
def tools():
    return render_template('tools.html', active='tools')


@app.route('/tools/spreadsheet', methods=['GET', 'POST'])
def spreadsheet_tool():
    """Parse a Google Sheets URL into products.json. Wraps parse_spreadsheet.py."""
    if request.method == 'POST':
        url = (request.form.get('url') or '').strip()
        target = (request.form.get('target') or 'kai').strip()
        cny_rate = request.form.get('cny_rate') or '6.5'
        if not url:
            return render_template('tool_spreadsheet.html', error='Sheet URL required',
                                   sites=load_sites(), active='tools')
        target_path = CLIENTS / target / 'static' / 'products.json'
        target_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            result = subprocess.run(
                [sys.executable, str(ROOT / 'scripts' / 'parse_spreadsheet.py'),
                 url, str(target_path), '--cny-rate', cny_rate],
                capture_output=True, text=True, timeout=120,
            )
            output = (result.stdout or '') + '\n' + (result.stderr or '')
        except subprocess.TimeoutExpired:
            output = 'Timed out after 120s.'
        return render_template('tool_spreadsheet.html', output=output, target=target,
                               sites=load_sites(), active='tools')
    return render_template('tool_spreadsheet.html', sites=load_sites(), active='tools')


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    print(f'\n  Master Admin running at http://127.0.0.1:{port}')
    print(f'  Password: {MASTER_PASSWORD}\n')
    app.run(host='0.0.0.0', port=port, debug=True)
