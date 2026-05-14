"""
Smart spreadsheet parser for fashion finds spreadsheets.
Detects columns automatically: product name, link, price (CNY/USD), image.

Usage:
    python parse_spreadsheet.py <google_sheets_url> <output_products.json> [--cny-rate 6.5]
"""
import sys, os, json, re, secrets, requests, argparse
from urllib.parse import unquote, urlparse, parse_qs

# Make the per-site tag_utils.py importable. We try the destination site folder
# first (where parse_spreadsheet.py is invoked against), then fall back to the
# Kai Finds template.
_THIS = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_THIS, '..', '..', 'kai'))
try:
    from tag_utils import generate_tags
except ImportError:
    def generate_tags(name, category=''):
        return (category or '').lower()


def download_spreadsheet(url, dest='_spreadsheet.xlsx'):
    """Convert any Google Sheets URL to XLSX export URL and download."""
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if not m:
        raise ValueError(f'Not a valid Google Sheets URL: {url}')
    sheet_id = m.group(1)
    xlsx_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'
    print(f'Downloading from {xlsx_url}...', flush=True)
    r = requests.get(xlsx_url, timeout=60)
    if r.status_code != 200:
        raise ValueError(f'Download failed (status {r.status_code}). Make sure the sheet is shared as "Anyone with the link can view".')
    with open(dest, 'wb') as f:
        f.write(r.content)
    print(f'Saved {len(r.content)} bytes', flush=True)
    return dest


def detect_columns(rows, sample_size=20):
    """Look at the first N rows and figure out which column is what."""
    sample = rows[:sample_size]
    n_cols = max((len(r) for r in sample), default=0)
    if n_cols == 0:
        return {}

    scores = {i: {'name': 0, 'link': 0, 'price': 0, 'image': 0, 'cny': 0, 'usd': 0} for i in range(n_cols)}

    for row in sample:
        for i, cell in enumerate(row):
            if i >= n_cols:
                continue
            val = (cell.get('value') or '') if isinstance(cell, dict) else (cell or '')
            val_str = str(val).strip()
            link = cell.get('hyperlink', '') if isinstance(cell, dict) else ''

            # Link column: has hyperlinks or http URLs
            if link or (val_str.startswith('http') and 'item' in val_str):
                scores[i]['link'] += 2
            if val_str.upper() == 'LINK':
                scores[i]['link'] += 1

            # Image column: =IMAGE() formula
            if 'IMAGE(' in val_str.upper() or (val_str.startswith('http') and any(ext in val_str.lower() for ext in ['.jpg', '.png', '.webp', '.jpeg'])):
                scores[i]['image'] += 2

            # Price columns
            if '$' in val_str:
                scores[i]['usd'] += 2
            if '¥' in val_str or 'CNY' in val_str.upper():
                scores[i]['cny'] += 2
            # Numeric-only could be CNY
            try:
                num = float(val_str.replace(',', '').replace('¥', '').replace('$', '').strip())
                if 5 < num < 50000:
                    scores[i]['price'] += 1
                    if num > 100:  # likely CNY if high
                        scores[i]['cny'] += 0.5
                    else:
                        scores[i]['usd'] += 0.5
            except (ValueError, AttributeError):
                pass

            # Name column: long text, not a URL
            if 10 < len(val_str) < 100 and not val_str.startswith('http') and not val_str.startswith('='):
                scores[i]['name'] += 0.5

    # Pick best column for each role
    detected = {}
    for role in ['name', 'link', 'cny', 'usd', 'image']:
        best_col = max(scores.keys(), key=lambda c: scores[c][role], default=None)
        if best_col is not None and scores[best_col][role] > 0:
            detected[role] = best_col

    return detected


def parse_xlsx(xlsx_path, cny_rate=6.5):
    """Parse an XLSX file and return products."""
    try:
        import openpyxl
    except ImportError:
        os.system(f'{sys.executable} -m pip install openpyxl -q')
        import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    all_products = []
    seen_keys = set()

    # Categorize sheets by name keywords
    sheet_cat_map = {
        'trending': 'trending', 'latest': 'trending', 'popular': 'trending', 'best': 'trending',
        'shoe': 'shoes', 'sneaker': 'shoes', 'footwear': 'shoes',
        'shirt': 'shirts', 'tee': 'shirts', 't-shirt': 'shirts', 'top': 'shirts', 'polo': 'shirts',
        'hoodie': 'hoodies', 'sweat': 'hoodies', 'crewneck': 'hoodies',
        'pant': 'pants', 'jean': 'pants', 'short': 'pants', 'trouser': 'pants', 'jogger': 'pants',
        'jacket': 'jackets', 'coat': 'jackets', 'puffer': 'jackets', 'outerwear': 'jackets',
        'access': 'accessories', 'bag': 'accessories', 'wallet': 'accessories', 'belt': 'accessories', 'watch': 'accessories', 'hat': 'accessories',
        'electronic': 'tech', 'tech': 'tech', 'gadget': 'tech',
        'women': 'womens',
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cat = 'trending'
        for kw, c in sheet_cat_map.items():
            if kw in sheet_name.lower():
                cat = c
                break

        print(f'  Sheet "{sheet_name}" → category "{cat}"', flush=True)

        # Build rows with cell data + hyperlinks
        rows = []
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            row_data = []
            for c in r:
                row_data.append({
                    'value': c.value,
                    'hyperlink': c.hyperlink.target if c.hyperlink else None,
                })
            rows.append(row_data)

        # Auto-detect columns
        cols = detect_columns(rows)
        if 'name' not in cols or 'link' not in cols:
            print(f'    Skipping (could not detect name+link columns)', flush=True)
            continue
        print(f'    Detected columns: {cols}', flush=True)

        # Extract products
        count = 0
        for row in rows:
            if len(row) <= max(cols.values(), default=0):
                continue

            name = str(row[cols['name']]['value'] or '').strip()
            if not name or len(name) < 2 or name.startswith('='):
                continue
            skip = ['JOYAGOO', 'KAKOBUY', 'SIGNUP', 'EXCHANGE', 'EURO', 'POPULAR', 'TELEGRAM',
                    'PRODUCT', 'TIKTOK', 'US DOLLAR', 'EURO EXCHANGE', 'PLEASE DO NOT', 'LATEST FINDS']
            if any(x.lower() in name.lower() for x in skip):
                continue

            # Get URL
            url = ''
            link_cell = row[cols['link']]
            if link_cell.get('hyperlink'):
                url = str(link_cell['hyperlink'])
            elif isinstance(link_cell['value'], str) and link_cell['value'].startswith('http'):
                url = link_cell['value']
            if not url:
                continue

            # Dedupe by name+url
            key = f'{name.lower()}|{url}'
            if key in seen_keys:
                continue
            seen_keys.add(key)

            # Price: prefer USD, else convert from CNY
            price_usd = ''
            if 'usd' in cols:
                v = str(row[cols['usd']]['value'] or '')
                if '$' in v:
                    m = re.search(r'(\d+\.?\d*)', v.replace(',', ''))
                    if m:
                        price_usd = m.group(1)
            if not price_usd and 'cny' in cols:
                try:
                    cny = float(str(row[cols['cny']]['value']).replace(',', '').strip())
                    if 1 < cny < 50000:
                        price_usd = f'{cny / cny_rate:.2f}'
                except (ValueError, TypeError, AttributeError):
                    pass
            if not price_usd:
                continue
            try:
                price_float = float(price_usd)
            except ValueError:
                continue
            if price_float < 0.5 or price_float > 2000:
                continue

            # Image
            image = ''
            if 'image' in cols and row[cols['image']]['value']:
                v = str(row[cols['image']]['value'])
                m = re.search(r'https?://[^\s\)"\']+', v)
                if m:
                    image = m.group(0).rstrip('"\'.)')

            # Clean name (strip emoji)
            name = re.sub(r'[\U0001F300-\U0001F9FF\U0001FA00-\U0001FAFF☀-⛿✀-➿]', '', name).strip()
            if len(name) < 2:
                continue

            clean_name = name[:80]
            all_products.append({
                'id': f'p{secrets.token_hex(4)}',
                'name': clean_name,
                'category': cat,
                'price': f'{price_float:.2f}',
                'price_numeric': price_float,
                'url': url,
                'image': image,
                'seller': '',
                'batch': '',
                'retail_price': '',
                'tags': generate_tags(clean_name, cat),
            })
            count += 1
        print(f'    Extracted {count} products', flush=True)

    return all_products


def main():
    p = argparse.ArgumentParser()
    p.add_argument('url', help='Google Sheets URL or path to local XLSX')
    p.add_argument('output', help='Output products.json path')
    p.add_argument('--cny-rate', type=float, default=6.5, help='CNY to USD divider (default 6.5)')
    args = p.parse_args()

    # Download if URL
    if args.url.startswith('http'):
        xlsx = download_spreadsheet(args.url)
    else:
        xlsx = args.url

    print(f'\nParsing spreadsheet...', flush=True)
    products = parse_xlsx(xlsx, cny_rate=args.cny_rate)

    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    from collections import Counter
    cats = Counter(p['category'] for p in products)
    print(f'\n✓ Total: {len(products)} products')
    for c, n in cats.most_common():
        print(f'  {c}: {n}')
    print(f'\nSaved to {args.output}')


if __name__ == '__main__':
    main()
