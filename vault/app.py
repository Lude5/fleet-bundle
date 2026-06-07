import os
import json
import secrets
from flask import Flask, render_template, request, jsonify, redirect, session, url_for, send_file
from flask_cors import CORS
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
CORS(app)

# === SITE CONFIG (all customizable via env vars) ===
# Every visible string and toggle on the Minimal template can be overridden via
# Render env vars. No code edits needed to deploy a new client — just paste env
# vars into the Render dashboard.

def _bool(name, default='1'):
    """Parse '1'/'true'/'yes' as True, anything else as False."""
    return os.environ.get(name, default).strip().lower() in ('1', 'true', 'yes', 'on')


SITE_CONFIG = {
    # ---- Brand ----
    'name':              os.environ.get('SITE_NAME', 'VAULT'),
    'name_part1':        os.environ.get('SITE_NAME_PART1', 'VAULT'),
    'name_part2':        os.environ.get('SITE_NAME_PART2', '.'),
    'domain':            os.environ.get('SITE_DOMAIN', 'vault.com'),
    'tagline':           os.environ.get('TAGLINE', 'Curated luxury. Quietly exceptional.'),
    'product_count_label': os.environ.get('PRODUCT_COUNT_LABEL', '100+'),

    # ---- Agent / affiliate ----
    'agent_name':        os.environ.get('AGENT_NAME', 'KakoBuy'),
    'agent_domain':      os.environ.get('AGENT_DOMAIN', 'kakobuy.com'),
    'agent_signup_url':  os.environ.get('AGENT_SIGNUP_URL', 'https://www.kakobuy.com/register?affcode=vault'),
    'agent_product_url': os.environ.get('AGENT_PRODUCT_URL', 'https://www.kakobuy.com/item/details'),
    'affiliate_code':    os.environ.get('AFFILIATE_CODE', 'vault'),
    'coupon_amount':     os.environ.get('COUPON_AMOUNT', '500'),

    # ---- Colors (CSS variables — preview-mode.js can override these live) ----
    # VAULT: warm ivory canvas (whitespace), near-black ink, antique gold accent.
    'brand_color':        os.environ.get('BRAND_COLOR', '#b08d4f'),         # --accent (gold)
    'brand_color_shadow': os.environ.get('BRAND_COLOR_SHADOW', '#8a6d3a'),  # darker gold hover
    'bg_color':           os.environ.get('BG_COLOR', '#fbfaf7'),            # --bg (warm ivory)
    'text_color':         os.environ.get('TEXT_COLOR', '#12110f'),          # --text (warm near-black)
    'text_secondary':     os.environ.get('TEXT_SECONDARY', '#6f6a60'),      # --text-secondary
    'text_muted':         os.environ.get('TEXT_MUTED', '#a8a294'),          # --text-muted
    'surface_color':      os.environ.get('SURFACE_COLOR', '#f3f0e9'),       # --surface
    'border_color':       os.environ.get('BORDER_COLOR', '#e6e1d5'),        # --border-subtle (warm hairline)

    # ---- Hero copy ----
    'hero_eyebrow':      os.environ.get('HERO_EYEBROW', 'Maison · Est. 2026'),  # custom eyebrow; empty = auto from name+count
    'hero_headline':     os.environ.get('HERO_HEADLINE', 'Quiet<br>luxury.'),
    'hero_meta':         os.environ.get('HERO_META', ''),                   # subtext under headline; empty = auto from tagline
    'hero_cta_primary':  os.environ.get('HERO_CTA_PRIMARY', 'Browse all'),
    'hero_cta_signup':   os.environ.get('HERO_CTA_SIGNUP', ''),             # empty = auto "Sign up · $X"

    # ---- Section labels ----
    'label_latest':      os.environ.get('LABEL_LATEST', 'Latest'),
    'label_featured':    os.environ.get('LABEL_FEATURED', 'Best selling'),
    'label_index':       os.environ.get('LABEL_INDEX', 'Index'),
    'label_view_all':    os.environ.get('LABEL_VIEW_ALL', 'View all →'),

    # ---- Mid-CTA bar ----
    'mid_cta_big':       os.environ.get('MID_CTA_BIG', ''),                 # empty = auto "$X in welcome coupons"
    'mid_cta_small':     os.environ.get('MID_CTA_SMALL', ''),               # empty = auto "Sign up to <agent> · lowest shipping rates"

    # ---- Index strip (marquee) ----
    'index_strip_words': os.environ.get(
        'INDEX_STRIP_WORDS',
        'Curated · {agent} · Weidian / Taobao / 1688 · Lowest shipping · Black · White · Gold · Quietly exceptional'
    ),

    # ---- About block ----
    'about_headline':    os.environ.get('ABOUT_HEADLINE', ''),              # empty = auto
    'about_body_1':      os.environ.get('ABOUT_BODY_1', ''),                # empty = auto
    'about_body_2':      os.environ.get('ABOUT_BODY_2', ''),                # empty = auto

    # ---- Closing CTA ----
    'closing_eyebrow':   os.environ.get('CLOSING_EYEBROW', 'Welcome offer'),
    'closing_headline':  os.environ.get('CLOSING_HEADLINE', ''),            # empty = auto "$X in coupons."
    'closing_body':      os.environ.get('CLOSING_BODY', ''),                # empty = auto

    # ---- Popup ----
    'popup_eyebrow':     os.environ.get('POPUP_EYEBROW', 'Welcome offer'),
    'popup_headline':    os.environ.get('POPUP_HEADLINE', ''),              # empty = auto "$X in coupons."
    'popup_sub':         os.environ.get('POPUP_SUB', ''),                   # empty = auto
    'popup_cta':         os.environ.get('POPUP_CTA', ''),                   # empty = auto "Sign up to <agent>"

    # ---- Footer ----
    'footer_desc':       os.environ.get('FOOTER_DESC', 'Curated {agent} finds. Black, white, gold. Nothing else.'),
    'footer_tagline':    os.environ.get('FOOTER_TAGLINE', 'Quietly exceptional'),

    # ---- Feature toggles (on/off) ----
    'enable_popup':       _bool('ENABLE_POPUP', '1'),
    'enable_sales_notif': _bool('ENABLE_SALES_NOTIF', '1'),
    'enable_side_widget': _bool('ENABLE_SIDE_WIDGET', '1'),
    'enable_index_strip': _bool('ENABLE_INDEX_STRIP', '1'),
    'enable_conveyor':    _bool('ENABLE_CONVEYOR', '1'),
    'enable_about':       _bool('ENABLE_ABOUT', '1'),
    'enable_closing_cta': _bool('ENABLE_CLOSING_CTA', '1'),
    'enable_mid_cta':     _bool('ENABLE_MID_CTA', '1'),
    'enable_currency':    _bool('ENABLE_CURRENCY', '1'),

    # ---- Misc ----
    'meta_pixel_id':     os.environ.get('META_PIXEL_ID', ''),
    'discord_url':       os.environ.get('DISCORD_URL', ''),
    'popup_delay_ms':    int(os.environ.get('POPUP_DELAY_MS', '2000')),     # how long after page load to show popup
}

# Resolve {agent} placeholders in copy that references the agent name
def _resolve(s):
    return s.replace('{agent}', SITE_CONFIG['agent_name']) if isinstance(s, str) else s


for _k in ('index_strip_words', 'footer_desc', 'tagline'):
    SITE_CONFIG[_k] = _resolve(SITE_CONFIG[_k])

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'changeme123')
ADMIN_API_TOKEN = os.environ.get('ADMIN_API_TOKEN', '')  # for cross-site master-admin calls
DATA_DIR = os.environ.get('VAULT_DATA_DIR') or ('/data' if os.path.exists('/data') else os.path.join(os.path.dirname(__file__), 'data'))
os.makedirs(DATA_DIR, exist_ok=True)

try:
    try:
        from .database import (
            init_db, get_products, get_product, add_product, add_products_bulk,
            update_product, delete_product, search_products, get_categories, add_category,
            record_click, get_analytics, backup_database, check_auto_backup,
            set_featured, move_category, reorder_products, get_listing_variants
        )
    except ImportError:
        from database import (
            init_db, get_products, get_product, add_product, add_products_bulk,
            update_product, delete_product, search_products, get_categories, add_category,
            record_click, get_analytics, backup_database, check_auto_backup,
            set_featured, move_category, reorder_products, get_listing_variants
        )
    init_db()
    check_auto_backup()

    if not get_products():
        products_file = os.path.join(os.path.dirname(__file__), 'static', 'products.json')
        if os.path.exists(products_file):
            with open(products_file, 'r', encoding='utf-8') as _f:
                _products = json.load(_f)
            add_products_bulk(_products)
            print(f"Loaded {len(_products)} products")

        CATS = [
            {'slug': 'trending', 'name': 'Trending', 'sort_order': 0},
            {'slug': 'shoes', 'name': 'Shoes', 'sort_order': 1},
            {'slug': 'shirts', 'name': 'Shirts', 'sort_order': 2},
            {'slug': 'hoodies', 'name': 'Hoodies', 'sort_order': 3},
            {'slug': 'pants', 'name': 'Pants', 'sort_order': 4},
            {'slug': 'jackets', 'name': 'Jackets', 'sort_order': 5},
            {'slug': 'accessories', 'name': 'Accessories', 'sort_order': 6},
            {'slug': 'bags', 'name': 'Bags', 'sort_order': 7},
            {'slug': 'tech', 'name': 'Tech', 'sort_order': 8},
            {'slug': 'womens', 'name': 'Womens', 'sort_order': 9},
        ]
        for c in CATS:
            add_category(c['slug'], c['name'], '', '', c['sort_order'])
        print("Categories seeded")
except Exception as e:
    print(f"DB init warning: {e}")


def is_admin():
    return session.get('admin_logged_in', False)


def is_admin_api():
    """Either a logged-in admin session OR a valid X-Admin-Token header.

    Master admin calls cross-site APIs with the header version.
    """
    if session.get('admin_logged_in', False):
        return True
    token = request.headers.get('X-Admin-Token') or request.args.get('token')
    if ADMIN_API_TOKEN and token and token == ADMIN_API_TOKEN:
        return True
    return False


@app.context_processor
def inject_config():
    """Make SITE_CONFIG available in all templates as 'site'."""
    return {'site': SITE_CONFIG}


# --- Public Routes ---

@app.route('/')
def home():
    products = get_products()
    categories = get_categories()
    import random
    shuffled = list(products)
    random.shuffle(shuffled)
    return render_template('home.html',
        products=products,
        conveyor=shuffled[:40],
        hero_products=shuffled[:24],
        categories=categories)


@app.route('/shop')
def shop():
    category = request.args.get('category', '')
    sort = request.args.get('sort', 'newest')
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 40

    if q:
        all_products = search_products(q)
        if category:
            all_products = [p for p in all_products if p.get('category') == category]
    else:
        all_products = get_products(category if category else None)

    if sort == 'price_low':
        all_products.sort(key=lambda p: p.get('price_numeric', 0))
    elif sort == 'price_high':
        all_products.sort(key=lambda p: p.get('price_numeric', 0), reverse=True)

    total = len(all_products)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    products = all_products[(page - 1) * per_page : page * per_page]

    return render_template('shop.html', products=products, categories=get_categories(),
        current_category=category, current_sort=sort, search_query=q,
        page=page, total_pages=total_pages, total=total)


@app.route('/link-converter')
def link_converter():
    return render_template('link_converter.html')


@app.route('/tutorial')
def tutorial():
    return render_template('tutorial.html')


@app.route('/spreadsheet')
def spreadsheet():
    products = get_products()
    return render_template('spreadsheet.html', products=products, categories=get_categories(), total=len(products))


@app.route('/privacy')
def privacy():
    return render_template('privacy.html')


@app.route('/terms')
def terms():
    return render_template('terms.html')


@app.route('/robots.txt')
def robots():
    domain = SITE_CONFIG['domain']
    return f"""User-agent: *
Allow: /
Sitemap: https://{domain}/sitemap.xml

User-agent: Googlebot
Allow: /
Crawl-delay: 1

User-agent: ChatGPT-User
Allow: /

User-agent: GPTBot
Allow: /

User-agent: OAI-SearchBot
Allow: /

User-agent: Bingbot
Allow: /
""", 200, {'Content-Type': 'text/plain'}


@app.route('/llms.txt')
def llms_txt():
    s = SITE_CONFIG
    return f"""# {s['name']}

> The #1 {s['agent_name']} finds website. Browse curated products at the best prices. Sign up to {s['agent_name']} for ${s['coupon_amount']} in welcome coupons.

## Pages

- [Home](https://{s['domain']}/): Browse all {s['agent_name']} products
- [Shop](https://{s['domain']}/shop): Search and filter products
- [Link Converter](https://{s['domain']}/link-converter): Convert any agent link to {s['agent_name']}
- [Spreadsheet](https://{s['domain']}/spreadsheet): Full product spreadsheet

## Keywords
{s['agent_name'].lower()}, {s['name'].lower()}, {s['agent_name'].lower()} products, {s['agent_name'].lower()} finds, {s['agent_name'].lower()} coupons, rep finds
""", 200, {'Content-Type': 'text/plain'}


@app.route('/sitemap.xml')
def sitemap():
    domain = SITE_CONFIG['domain']
    pages = ['/', '/shop', '/spreadsheet', '/link-converter', '/privacy', '/terms']
    for cat in get_categories():
        pages.append(f'/shop?category={cat["slug"]}')
    xml = '<?xml version="1.0" encoding="UTF-8"?>\n<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    for page in pages:
        xml += f'  <url><loc>https://{domain}{page}</loc><changefreq>daily</changefreq><priority>{"1.0" if page == "/" else "0.8"}</priority></url>\n'
    xml += '</urlset>'
    return xml, 200, {'Content-Type': 'application/xml'}


@app.route('/go/<product_id>')
def affiliate_redirect(product_id):
    product = get_product(product_id)
    if not product:
        return redirect(url_for('shop'))
    try:
        record_click({
            'product_id': product_id, 'product_name': product.get('name', ''),
            'category': product.get('category', ''), 'element_type': 'affiliate',
            'page': 'redirect', 'referrer': request.referrer or '',
            'user_ip': request.headers.get('X-Forwarded-For', request.remote_addr or ''),
            'user_agent': request.headers.get('User-Agent', ''), 'country': '',
        })
    except Exception:
        pass
    url = product.get('url', '')
    if url:
        from urllib.parse import quote
        agent_url = f"{SITE_CONFIG['agent_product_url']}?url={quote(url)}"
        if SITE_CONFIG['affiliate_code']:
            agent_url += f"&affcode={SITE_CONFIG['affiliate_code']}"
        return redirect(agent_url)
    return redirect(url_for('shop'))


# --- API ---

@app.route('/api/product/<pid>')
def api_product(pid):
    """Return a product with its listing variants and (stubbed) QC photos.
    Used by the product detail modal."""
    p = get_product(pid)
    if not p:
        return jsonify({'error': 'not found'}), 404
    variants = get_listing_variants(p)
    # If no real variants found (e.g. demo data has no listing_id), still
    # return the product itself as the sole variant.
    if not variants:
        variants = [p]
    # QC photos placeholder — will be wired to Kakobuy QC API later.
    qc_photos = []
    if p.get('qc_photos'):
        try:
            qc_photos = json.loads(p['qc_photos'])
        except (ValueError, TypeError):
            qc_photos = []
    return jsonify({
        'id': p['id'],
        'name': p['name'],
        'price': p['price'],
        'price_numeric': p.get('price_numeric', 0),
        'image': p['image'],
        'category': p.get('category', ''),
        'seller': p.get('seller', ''),
        'batch': p.get('batch', ''),
        'retail_price': p.get('retail_price', ''),
        'tags': p.get('tags', ''),
        'weight': p.get('weight', ''),
        'quality': p.get('quality', ''),
        'sales': p.get('sales', 0),
        'go_url': f'/go/{pid}',
        'variants': [
            {
                'id': v['id'],
                'name': v['name'],
                'image': v['image'],
                'price': v['price'],
                'price_numeric': v.get('price_numeric', 0),
            } for v in variants
        ],
        'qc_photos': qc_photos,
    })


@app.route('/api/products')
def api_products():
    limit = request.args.get('limit', 50, type=int)
    products = get_products()
    import random
    sampled = random.sample(products, min(limit, len(products)))
    return jsonify([{'id': p['id'], 'name': p['name'], 'image': p['image'], 'price': p['price'], 'category': p['category']} for p in sampled])


@app.route('/api/click', methods=['POST'])
def api_click():
    data = request.get_json(silent=True) or {}
    data['user_ip'] = request.headers.get('X-Forwarded-For', request.remote_addr or '')
    data['user_agent'] = request.headers.get('User-Agent', '')
    try:
        record_click(data)
    except Exception:
        pass
    return jsonify({'ok': True})


@app.route('/api/track-pageview', methods=['POST'])
def api_pageview():
    data = request.get_json(silent=True) or {}
    data['element_type'] = 'pageview'
    data['user_ip'] = request.headers.get('X-Forwarded-For', request.remote_addr or '')
    data['user_agent'] = request.headers.get('User-Agent', '')
    try:
        record_click(data)
    except Exception:
        pass
    return jsonify({'ok': True})


# --- Admin ---

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if request.form.get('password') == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        return render_template('admin_login.html', error='Wrong password')
    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('home'))


@app.route('/admin')
def admin_dashboard():
    if not is_admin():
        return redirect(url_for('admin_login'))
    stats = get_analytics(30)
    products = get_products()
    categories = get_categories()
    featured_count = sum(1 for p in products if p.get('featured'))
    cats_with_counts = []
    for c in categories:
        n = sum(1 for p in products if p.get('category') == c.get('slug'))
        cats_with_counts.append({**c, 'count': n})
    recent_products = sorted(products, key=lambda p: p.get('updated_at') or '', reverse=True)[:8]
    return render_template(
        'admin_dashboard.html',
        stats=stats,
        products=products,
        categories=cats_with_counts,
        featured_count=featured_count,
        recent_products=recent_products,
        total_products=len(products),
    )


@app.route('/admin/products/add', methods=['POST'])
def admin_add_product():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    if not data.get('name'):
        return jsonify({'error': 'Name required'}), 400
    data['id'] = data.get('id', f"p{secrets.token_hex(4)}")
    if not data.get('tags'):
        try:
            from .tag_utils import generate_tags
        except ImportError:
            from tag_utils import generate_tags
        data['tags'] = generate_tags(data['name'], data.get('category', ''))
    add_product(data)
    return jsonify({'ok': True, 'id': data['id']})


@app.route('/admin/products/update/<pid>', methods=['POST'])
def admin_update_product(pid):
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({'error': 'No data'}), 400
    if 'price' in data:
        try:
            data['price_numeric'] = float(data['price'] or 0)
        except ValueError:
            data['price_numeric'] = 0
    update_product(pid, data)
    return jsonify({'ok': True})


@app.route('/admin/products/bulk', methods=['POST'])
def admin_bulk():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    products = data.get('products', [])
    if not products:
        return jsonify({'error': 'No products'}), 400
    try:
        from .tag_utils import generate_tags
    except ImportError:
        from tag_utils import generate_tags
    for p in products:
        if not p.get('id'):
            p['id'] = f"p{secrets.token_hex(4)}"
        if not p.get('tags'):
            p['tags'] = generate_tags(p.get('name', ''), p.get('category', ''))
    add_products_bulk(products)
    return jsonify({'ok': True, 'count': len(products)})


@app.route('/admin/products/delete/<pid>', methods=['DELETE'])
def admin_delete(pid):
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    delete_product(pid)
    return jsonify({'ok': True})


@app.route('/admin/products/delete-batch', methods=['POST'])
def admin_delete_batch():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    for pid in ids:
        delete_product(pid)
    return jsonify({'ok': True, 'count': len(ids)})


@app.route('/admin/products/feature', methods=['POST'])
def admin_feature():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    featured = bool(data.get('featured', True))
    n = set_featured(ids, featured)
    return jsonify({'ok': True, 'count': n, 'featured': featured})


@app.route('/admin/products/move-category', methods=['POST'])
def admin_move_category():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    category = (data.get('category') or '').strip()
    if not category:
        return jsonify({'error': 'Category required'}), 400
    n = move_category(ids, category)
    return jsonify({'ok': True, 'count': n, 'category': category})


@app.route('/admin/products/reorder', methods=['POST'])
def admin_reorder():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    n = reorder_products(ids)
    return jsonify({'ok': True, 'count': n})


@app.route('/admin/products')
def admin_products():
    if not is_admin():
        return redirect(url_for('admin_login'))
    return render_template('admin_products.html', products=get_products(), categories=get_categories())


@app.route('/admin/analytics')
def admin_analytics():
    if not is_admin():
        return redirect(url_for('admin_login'))
    from datetime import datetime, timedelta

    range_key = request.args.get('range', '30d')
    now = datetime.now()
    since = None
    until = now

    if range_key == 'today':
        since = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif range_key == 'hour':
        since = now - timedelta(hours=1)
    elif range_key == '24h':
        since = now - timedelta(hours=24)
    elif range_key == '7d':
        since = now - timedelta(days=7)
    elif range_key == '30d':
        since = now - timedelta(days=30)
    elif range_key == '90d':
        since = now - timedelta(days=90)
    elif range_key == 'mtd':  # Month-to-date: 1st of this month → now
        since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif range_key == 'ytd':  # Year-to-date: Jan 1 → now
        since = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    elif range_key == '1y':
        since = now - timedelta(days=365)
    elif range_key == 'custom':
        f = request.args.get('from')
        t = request.args.get('to')
        if f:
            try: since = datetime.fromisoformat(f)
            except ValueError: pass
        if t:
            try:
                # Treat "to" as end-of-day so 2026-05-11 includes all of that day
                until_dt = datetime.fromisoformat(t)
                until = until_dt.replace(hour=23, minute=59, second=59, microsecond=999999) if until_dt.hour == 0 and until_dt.minute == 0 else until_dt
            except ValueError: pass

    if since is None:
        since = now - timedelta(days=30)
        range_key = '30d'

    stats = get_analytics(since=since.isoformat(), until=until.isoformat())
    return render_template(
        'admin_analytics.html',
        stats=stats,
        range_key=range_key,
        range_since=since.strftime('%Y-%m-%d'),
        range_until=until.strftime('%Y-%m-%d'),
        # Hours-in-range for "Daily Avg" math
        span_seconds=(until - since).total_seconds(),
    )


@app.route('/admin/analytics/export')
def admin_analytics_export():
    """Export one of the analytics breakdowns as CSV."""
    if not is_admin():
        return redirect(url_for('admin_login'))
    from datetime import datetime, timedelta
    from io import StringIO
    import csv

    kind = request.args.get('kind', 'products')
    range_key = request.args.get('range', '30d')
    now = datetime.now()
    until = now
    if range_key == 'today':   since = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif range_key == 'hour':  since = now - timedelta(hours=1)
    elif range_key == '24h':   since = now - timedelta(hours=24)
    elif range_key == '7d':    since = now - timedelta(days=7)
    elif range_key == '30d':   since = now - timedelta(days=30)
    elif range_key == '90d':   since = now - timedelta(days=90)
    elif range_key == 'mtd':   since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif range_key == 'ytd':   since = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    elif range_key == '1y':    since = now - timedelta(days=365)
    elif range_key == 'custom':
        try: since = datetime.fromisoformat(request.args.get('from') or '')
        except (ValueError, TypeError): since = now - timedelta(days=30)
        try: until = datetime.fromisoformat(request.args.get('to') or '').replace(hour=23, minute=59, second=59)
        except (ValueError, TypeError): pass
    else:                      since = now - timedelta(days=30)

    stats = get_analytics(since=since.isoformat(), until=until.isoformat())
    fmt = (request.args.get('format') or 'csv').lower()

    # Build rows (kind → (headers, list-of-rows))
    if kind == 'products':
        headers = ['Rank', 'Product', 'Clicks']
        rows = [[i, p['product_name'], p['clicks']] for i, p in enumerate(stats['top_products'], 1)]
    elif kind == 'categories':
        headers = ['Rank', 'Category', 'Clicks']
        rows = [[i, c['category'], c['clicks']] for i, c in enumerate(stats['top_categories'], 1)]
    elif kind == 'pages':
        headers = ['Rank', 'Page', 'Views']
        rows = [[i, p['page'], p['views']] for i, p in enumerate(stats['top_pages'], 1)]
    elif kind == 'types':
        headers = ['Rank', 'Element Type', 'Clicks']
        rows = [[i, e['element_type'] or 'unknown', e['clicks']] for i, e in enumerate(stats['element_types'], 1)]
    elif kind == 'daily':
        headers = ['Date', 'Clicks', 'Visitors']
        rows = [[d['day'], d['clicks'], d['visitors']] for d in stats['daily']]
    else:
        return jsonify({'error': 'unknown kind'}), 400

    base_name = f'{SITE_CONFIG.get("name","kai").lower().replace(" ","-")}-{kind}-{range_key}-{since.strftime("%Y%m%d")}'

    from flask import make_response
    if fmt in ('xlsx', 'excel'):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return jsonify({'error': 'openpyxl not installed on server — pip install openpyxl'}), 500
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = kind.title()
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='06B6D4')
            cell.alignment = Alignment(horizontal='left')
        for r in rows:
            ws.append(r)
        # Auto-size columns
        for col_idx, h in enumerate(headers, 1):
            longest = max([len(str(h))] + [len(str(r[col_idx-1])) for r in rows], default=10)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(longest + 4, 60)
        ws.freeze_panes = 'A2'
        buf = BytesIO()
        wb.save(buf)
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="{base_name}.xlsx"'
        return resp

    # CSV (default)
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = f'attachment; filename="{base_name}.csv"'
    return resp


@app.route('/admin/categories/add', methods=['POST'])
def admin_add_category():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    if not data.get('slug') or not data.get('name'):
        return jsonify({'error': 'Slug and name required'}), 400
    add_category(data['slug'], data['name'], data.get('icon', ''), data.get('description', ''), data.get('sort_order', 0))
    return jsonify({'ok': True})


@app.route('/admin/backup', methods=['POST'])
def admin_backup():
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify({'ok': True, 'path': backup_database()})


# ===========================================================================
# Cross-site API (used by the master admin) — token-auth alternative.
# All routes accept either an admin session OR an X-Admin-Token header.
# ===========================================================================

@app.route('/admin/api/ping')
def api_ping():
    """Simple liveness check that also validates the token. Public side."""
    has_token = bool(ADMIN_API_TOKEN)
    token_ok = is_admin_api()
    return jsonify({
        'ok': True,
        'site': SITE_CONFIG.get('name'),
        'agent': SITE_CONFIG.get('agent_name'),
        'token_required': has_token,
        'token_valid': token_ok,
    })


@app.route('/admin/api/stats')
def api_stats():
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    from datetime import datetime, timedelta
    days = request.args.get('days', 30, type=int)
    since = (datetime.now() - timedelta(days=days)).isoformat()
    stats = get_analytics(days=days)
    products = get_products()
    return jsonify({
        'site': SITE_CONFIG.get('name'),
        'agent': SITE_CONFIG.get('agent_name'),
        'total_products': len(products),
        'featured_count': sum(1 for p in products if p.get('featured')),
        'categories': len(get_categories()),
        'total_clicks': stats['total_clicks'],
        'unique_visitors': stats['unique_visitors'],
        'signup_clicks': stats['signup_clicks'],
        'top_products': stats['top_products'],
        'top_categories': stats['top_categories'],
        'daily': stats['daily'],
        'days': days,
        'since': since,
    })


@app.route('/admin/api/products')
def api_admin_products():
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify({
        'site': SITE_CONFIG.get('name'),
        'count': None,
        'products': get_products(),
    })


@app.route('/admin/api/products', methods=['POST'])
def api_admin_add_product():
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    if not data.get('name'):
        return jsonify({'error': 'Name required'}), 400
    if not data.get('id'):
        data['id'] = f"p{secrets.token_hex(4)}"
    if not data.get('tags'):
        try:
            try:
                from .tag_utils import generate_tags
            except ImportError:
                from tag_utils import generate_tags
            data['tags'] = generate_tags(data['name'], data.get('category', ''))
        except ImportError:
            pass
    add_product(data)
    return jsonify({'ok': True, 'id': data['id']})


@app.route('/admin/api/products/<pid>', methods=['PUT', 'PATCH'])
def api_admin_update_product(pid):
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    if 'price' in data:
        try: data['price_numeric'] = float(data['price'] or 0)
        except ValueError: data['price_numeric'] = 0
    update_product(pid, data)
    return jsonify({'ok': True})


@app.route('/admin/api/products/<pid>', methods=['DELETE'])
def api_admin_delete_product(pid):
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    delete_product(pid)
    return jsonify({'ok': True})


@app.route('/admin/api/config')
def api_admin_config():
    if not is_admin_api():
        return jsonify({'error': 'Unauthorized'}), 401
    # Don't leak secrets — strip token-ish keys
    safe = {k: v for k, v in SITE_CONFIG.items() if 'token' not in k.lower() and 'secret' not in k.lower()}
    return jsonify(safe)


@app.route('/admin/backup/download')
def admin_download_backup():
    if not is_admin():
        return redirect(url_for('admin_login'))
    return send_file(backup_database(), as_attachment=True)


@app.errorhandler(404)
def not_found(e):
    return redirect(url_for('home'))


if __name__ == '__main__':
    os.makedirs(DATA_DIR, exist_ok=True)
    app.run(debug=True, port=int(os.environ.get('PORT', 5012)))
